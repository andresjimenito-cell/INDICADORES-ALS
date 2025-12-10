"""
Módulo para extraer y comparar la sección 'OFERTA ECONOMICA' de los archivos Excel
Funciones principales:
- parse_offer_sheet(fileobj): devuelve DataFrame con filas de items y columnas para cada alternativa
- aggregate_offers(uploaded_files): recorre una lista de objetos file-like (Streamlit) y devuelve DataFrame comparativa

Requiere: pandas, openpyxl
"""
from typing import List
import openpyxl
import pandas as pd


def _find_oferta_sheet_name(wb):
    names = wb.sheetnames
    oferta = next((s for s in names if s.lower().strip() in ['oferta economica','oferta_economica','ofertaeconomica','oferta']), None)
    return oferta


def parse_offer_sheet(fileobj) -> pd.DataFrame:
    """Lee la hoja 'OFERTA ECONOMICA' de un archivo Excel (file-like) y devuelve
    un DataFrame con los items y valores para Full, Rrigo y Alternativa.

    Columnas de salida:
    - section: sección (Sistema fondo, Cable, Y-TOOL, Servicios)
    - equipo: descripción (col C)
    - codigo: columna D
    - comentarios: columna E
    - cantidad: columna F
    - full_unit, full_total (G,H)
    - rrigo_pct, rrigo_estado, rrigo_unit, rrigo_total (I,L)
    - alt_pct, alt_estado, alt_unit, alt_total (N,Q)
    """
    wb = openpyxl.load_workbook(fileobj, data_only=True)
    sheet_name = _find_oferta_sheet_name(wb)
    if sheet_name is None:
        return pd.DataFrame()

    sh = wb[sheet_name]

    rows = []
    # mapeo de filas a secciones según el enunciado (1-based)
    section_map = {}
    for r in range(16, 46):
        section_map[r] = 'SISTEMA DE FONDO'
    for r in range(46, 55):
        section_map[r] = 'CABLE'
    for r in range(55, 62):
        section_map[r] = 'Y-TOOL'
    for r in range(62, 69):
        section_map[r] = 'SERVICIOS'

    for r in range(16, 69):
        # leer celdas por índice de columna (openpyxl 1-based)
        equipo_b = sh.cell(row=r, column=2).value  # B
        desc_c = sh.cell(row=r, column=3).value  # C
        # si ambas están vacías, saltar
        if (equipo_b is None or (isinstance(equipo_b, str) and equipo_b.strip() == '')) and (desc_c is None or (isinstance(desc_c, str) and desc_c.strip() == '')):
            continue

        def _val(col):
            try:
                return sh.cell(row=r, column=col).value
            except Exception:
                return None

        # preferir el valor en columna B como etiqueta de 'equipo'
        equipo_label = equipo_b if equipo_b not in [None, ''] else desc_c

        item = {
            'row': r,
            'section': section_map.get(r, 'OTROS'),
            'equipo': str(equipo_label).strip(),
            'codigo': _val(4),
            'comentarios': _val(5),
            'cantidad': _val(6),
            'full_unit': _val(7),
            'full_total': _val(8),
            'rrigo_pct': _val(9),
            'rrigo_estado': _val(10),
            'rrigo_unit': _val(11),
            'rrigo_total': _val(12),
            'alt_pct': _val(14),
            'alt_estado': _val(15),
            'alt_unit': _val(16),
            'alt_total': _val(17),
        }
        rows.append(item)

    df = pd.DataFrame(rows)
    # normalizar tipos numéricos
    for col in ['cantidad','full_unit','full_total','rrigo_unit','rrigo_total','alt_unit','alt_total']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    return df


def aggregate_offers(uploaded_files: List, provider_names: List[str]=None) -> pd.DataFrame:
    """Recibe una lista de objetos file-like (p.ej. Streamlit uploaded_files) y devuelve
    un DataFrame vertical con columnas: provider, section, equipo, cantidad, category, unit, total
    """
    rows = []
    for idx, f in enumerate(uploaded_files):
        try:
            f.seek(0)
        except Exception:
            pass
        provider = None
        if provider_names and idx < len(provider_names):
            provider = provider_names[idx]
        else:
            provider = getattr(f, 'name', f'file_{idx}')

        try:
            df = parse_offer_sheet(f)
        except Exception:
            df = pd.DataFrame()

        if df.empty:
            continue

        # añadir columnas de metadatos: provider y filename
        df = df.copy()
        df['provider'] = provider
        df['filename'] = getattr(f, 'name', f'file_{idx}')

        # asegurar tipos numéricos para totales/unidades
        for c in ['cantidad','full_unit','full_total','rrigo_pct','rrigo_unit','rrigo_total','alt_pct','alt_unit','alt_total']:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)

        # keep the row-level data (one row per item) so consumer can filter por categoría
        rows.append(df)

    if not rows:
        return pd.DataFrame()
    out = pd.concat(rows, ignore_index=True, sort=False)
    return out


def pivot_comparison(df: pd.DataFrame) -> pd.DataFrame:
    """Genera una tabla pivot donde cada proveedor tenga columnas con cantidad/unit/total por categoría.
    Devuelve un DataFrame listo para mostrar.
    """
    if df.empty:
        return df
    # normalizar nombres de proveedor
    df['provider'] = df['provider'].astype(str)
    # pivot para mostrar por (section, equipo, category, codigo) y provider columnas
    pivot = df.pivot_table(index=['section','equipo','codigo','category'], columns='provider', values=['cantidad','unit','total'], aggfunc='first')
    # aplanar columnas
    pivot.columns = [f"{val[0]}__{val[1]}" for val in pivot.columns]
    pivot = pivot.reset_index()
    return pivot
