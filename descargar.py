import pandas as pd
import matplotlib.pyplot as plt
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile

def exportar_excel_con_graficas(tablas: dict, graficas: dict) -> bytes:
    """
    Exporta un archivo Excel con varias hojas de tablas y gráficas.
    tablas: dict con nombre_hoja: DataFrame
    graficas: dict con nombre_hoja: figura matplotlib
    Devuelve el contenido binario del archivo Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = list(tablas.keys())[0]
    # Agregar la primera tabla
    for r in dataframe_to_rows(tablas[ws.title], index=False, header=True):
        ws.append(r)
    # Agregar el resto de tablas
    for nombre, df in list(tablas.items())[1:]:
        ws2 = wb.create_sheet(title=nombre)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws2.append(r)
    # Guardar gráficas como imágenes temporales y agregarlas
    for nombre, fig in graficas.items():
        img_path = tempfile.mktemp(suffix='.png')
        fig.savefig(img_path, bbox_inches='tight')
        ws_img = wb.create_sheet(title=f"Grafica_{nombre}")
        img = XLImage(img_path)
        img.width = 600
        img.height = 340
        ws_img.add_image(img, 'A1')
    # Guardar a bytes
    with io.BytesIO() as output:
        wb.save(output)
        return output.getvalue()

def exportar_resumen_performance(df_monthly) -> bytes:
    """
    Exporta el DataFrame mensual de performance a un Excel formateado.
    """
    if df_monthly is None or df_monthly.empty:
        return None
        
    df_export = df_monthly.copy()
    
    # Renombrar columnas para mejor legibilidad
    renames = {
        'Mes': 'Mes',
        'Pozos_Operativos': 'Pozos Operativos',
        'Pozos_ON': 'Pozos Activos',
        'Pozos_OFF': 'Pozos Inactivos',
        'RunLife_Promedio': 'Tiempo de Vida Promedio (días)',
        'RunLife_General': 'Tiempo de Vida Total (días)',
        'TMEF_Promedio': 'TMEF Promedio (días)',
        'RunLife_Efectivo': 'Tiempo de Vida Efectivo Total (días)',
        'RunLife_Efectivo_Fallados': 'Tiempo de Vida Efectivo Fallados (días)',
        'Indice_Falla_ON': 'Índice de Falla ON (%)',
        'Indice_Falla_ON_ALS': 'Índice de Falla ALS ON (%)'
    }
    df_export.rename(columns={k: v for k, v in renames.items() if k in df_export.columns}, inplace=True)
    
    # Convertir índices a porcentaje
    if 'Índice de Falla ON (%)' in df_export.columns:
        df_export['Índice de Falla ON (%)'] = df_export['Índice de Falla ON (%)'] * 100
    if 'Índice de Falla ALS ON (%)' in df_export.columns:
        df_export['Índice de Falla ALS ON (%)'] = df_export['Índice de Falla ALS ON (%)'] * 100
        
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Performance')
    return output.getvalue()

# Ejemplo de uso en Streamlit:
# from descargar import exportar_resumen_performance
# excel_bytes = exportar_resumen_performance(st.session_state['df_monthly_summary'])
# st.download_button('Descargar Excel', data=excel_bytes, ...)
