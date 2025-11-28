import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from io import BytesIO
import openpyxl
from PIL import Image
import base64
import re
import os
import theme as _theme_mod

# Import seguro desde el módulo `theme` con valores por defecto si falta algo
COLOR_PRINCIPAL = getattr(_theme_mod, 'COLOR_PRINCIPAL', '#00ff99')
COLOR_FONDO_OSCURO = getattr(_theme_mod, 'COLOR_FONDO_OSCURO', '#1a1a2e')
COLOR_FONDO_CONTENEDOR = getattr(_theme_mod, 'COLOR_FONDO_CONTENEDOR', 'rgba(25, 25, 40, 0.7)')
COLOR_SOMBRA = getattr(_theme_mod, 'COLOR_SOMBRA', 'rgba(0, 255, 153, 0.5)')
COLOR_FUENTE = getattr(_theme_mod, 'COLOR_FUENTE', '#FFFFFF')
get_color_sequence = getattr(_theme_mod, 'get_color_sequence', lambda mode=None: [
    COLOR_PRINCIPAL, '#58147b', '#FFDE31', '#5AFFDA'
])

def _default_plotly_layout(xa=None, ya=None, mode=None):
    xa = xa or COLOR_PRINCIPAL
    ya = ya or COLOR_PRINCIPAL
    return {
        'plot_bgcolor': COLOR_FONDO_OSCURO,
        'paper_bgcolor': COLOR_FONDO_OSCURO,
        'font_color': COLOR_FUENTE,
        'title_font_color': COLOR_PRINCIPAL,
        'xaxis': {'color': xa},
        'yaxis': {'color': ya}
    }

# ==================== CONFIGURACIÓN DE PÁGINA ====================
st.set_page_config(
    page_title="Evaluación Técnica ESP", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# Definición de Colores
GLOW_COLOR_INDIGO = "rgba(4, 0, 82, 1)" 
GLOW_COLOR_TRANSLUCENT = "rgba(4, 0, 82, 0.4)"
GLOW_COLOR_LIGHT = "rgba(4, 0, 82, 0.1)"
COLOR_PRINCIPAL = "#4A90E2"
COLOR_FONDO_CONTENEDOR_NEUTRO = "rgba(0, 0, 0, 0.1)" 
COLOR_TEXTO_CLARO = "#FFFFFF"
COLOR_TEXTO_OSCURO = "#333333"
COLOR_WARNING = "#FFA726"
COLOR_SUCCESS = "#66BB6A"
COLOR_ERROR = "#EF5350"

# ==================== ESTILOS CSS PERSONALIZADOS ====================
custom_css = f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap');
    
    html, body, [class*="css"] {{
        font-family: 'Inter', sans-serif;
    }}

    div[data-testid="stMetricLabel"] {{
        font-size: 1rem;
        font-weight: 600;
        color: var(--text-color) !important; 
    }}

    .main-header {{
        background: linear-gradient(135deg, {COLOR_PRINCIPAL} 10%, {GLOW_COLOR_INDIGO} 100%);
        padding: 2rem;
        border-radius:10px;
        color: {COLOR_TEXTO_CLARO};
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 10px 30px {GLOW_COLOR_TRANSLUCENT};
    }}
    
    .main-header h1 {{
        font-size: 2.5rem;
        font-weight: 700;
        margin-bottom: 0.5rem;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.5);
    }}
    
    .main-header p {{
        font-size: 1rem;
        opacity: 0.95;
        font-weight: 300;
    }}
    
    div[data-testid="stMetricValue"] {{
        font-size: 2rem;
        font-weight: 700;
        color: {COLOR_PRINCIPAL};
    }}

   /* MENSAJE DE SIN DATOS (DISCRETO Y MINIMALISTA) */
    /* =================================================== */
    .no-data-message {{
        background-color: var(--background-color); /* Usar el fondo neutro de Streamlit */
        border: 1px dashed rgba(150, 150, 150, 0.4); /* Borde punteado muy sutil */
        border-radius: 8px;
        padding: 10px; /* Reducido */
        text-align: center;
        margin: 0.5rem 0; /* Menos margen */
        color: #999999; /* Color de texto gris suave */
        box-shadow: none;
        transition: all 0.3s ease;
    }}
    
    .no-data-message h3 {{
        color: #AAAAAA; /* Título muy discreto */
        font-size: 0.9rem; /* MUY REDUCIDO */
        margin: 0 0 3px 0;
        font-weight: 600;
    }}
    
    .no-data-message p {{
        color: #BBBBBB;
        font-size: 0.75rem; /* El texto más pequeño posible */
        margin: 0;
        opacity: 0.8;
    }}

    /* TABLAS: CAJA GLOW + MÁXIMA SIMETRÍA */
    .stDataFrame, [data-testid="stHorizontalBlock"] {{
        background-color: {COLOR_FONDO_CONTENEDOR_NEUTRO}; 
        padding: 0.5rem !important;
        border-radius: 8px;
        border: 1px solid {GLOW_COLOR_LIGHT};
        box-shadow: 
            0 0 40px {GLOW_COLOR_TRANSLUCENT},
            0 0 60px {GLOW_COLOR_LIGHT},
            inset 0 0 20px {GLOW_COLOR_LIGHT};
        transition: all 0.3s ease;
        overflow: hidden !important;
    }}
    
    .stDataFrame:hover, [data-testid="stHorizontalBlock"]:hover {{
        transform: translateY(-2px);
        box-shadow: 
            0 0 50px {GLOW_COLOR_TRANSLUCENT},
            0 0 80px {GLOW_COLOR_INDIGO},
            inset 0 0 25px {GLOW_COLOR_LIGHT};
    }}

    /* GRÁFICAS: QUITAR CAJA EN TODOS LOS CASOS */
    .stPlotlyChart,
    [data-testid="stPlotlyChart"],
    .js-plotly-plot,
    .plotly,
    .plot-container,
    .modebar,
    .modebar-group {{
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        padding: 0 !important;
        margin: 0 !important;
        border-radius: 0 !important;
    }}

    /* TABLAS: SIN BORDES INTERNOS */
    .dataframe {{
        border: none !important;
        border-collapse: collapse !important;
        width: 100% !important;
        max-width: 100% !important;
        table-layout: auto !important;
        margin: 0 !important;
    }}

    .dataframe thead th {{
        background: linear-gradient(90deg, {COLOR_PRINCIPAL} 0%, {GLOW_COLOR_TRANSLUCENT} 100%) !important;
        color: {COLOR_TEXTO_CLARO} !important;
        font-weight: 600;
        padding: 10px 8px !important;
        border: none !important;
        text-align: center;
        font-size: 0.9rem;
    }}

    .dataframe tbody td {{
        padding: 8px !important;
        border: none !important;
        text-align: center;
        font-size: 0.85rem;
        background-color: transparent !important;
    }}

    .dataframe tbody tr:nth-child(even) {{
        background-color: {GLOW_COLOR_LIGHT} !important;
    }}

    .dataframe tbody tr:hover {{
        background-color: {GLOW_COLOR_TRANSLUCENT} !important;
        transition: background-color 0.3s ease !important;
    }}

    /* SCROLL INTERNO */
    .stDataFrame > div {{
        overflow-x: auto !important;
        overflow-y: hidden !important;
        scrollbar-width: thin;
        scrollbar-color: {COLOR_PRINCIPAL} transparent;
    }}

    .stDataFrame > div::-webkit-scrollbar {{
        height: 6px;
    }}
    .stDataFrame > div::-webkit-scrollbar-thumb {{
        background: {COLOR_PRINCIPAL};
        border-radius: 3px;
    }}

    /* EXPANDERS, BOTONES, INPUTS */
    .streamlit-expanderHeader {{
        background-color: {COLOR_FONDO_CONTENEDOR_NEUTRO};
        border-radius: 4px;
        font-weight: 600;
        color: {COLOR_PRINCIPAL};
        padding: 1rem;
        border: 1px solid {GLOW_COLOR_TRANSLUCENT};
        box-shadow: 
            inset 0 0 15px {GLOW_COLOR_LIGHT},
            0 0 30px {GLOW_COLOR_TRANSLUCENT},
            0 0 50px {GLOW_COLOR_LIGHT};
        transition: all 0.3s ease;
    }}
    
    .streamlit-expanderHeader:hover {{
        border-color: {COLOR_PRINCIPAL};
        box-shadow: 
            inset 0 0 20px {GLOW_COLOR_LIGHT},
            0 0 40px {GLOW_COLOR_TRANSLUCENT},
            0 0 70px {GLOW_COLOR_LIGHT};
    }}

    @keyframes pulse {{
        0%, 100% {{ opacity: 1; }}
        50% {{ opacity: 0.6; }}
    }}
    
    .stSpinner > div {{
        animation: pulse 1.5s ease-in-out infinite;
    }}
    
    button {{
        border: 2px solid {COLOR_PRINCIPAL} !important;
        box-shadow: 0 0 15px {GLOW_COLOR_LIGHT}, inset 0 0 5px {GLOW_COLOR_LIGHT} !important;
        transition: all 0.3s ease !important;
    }}
    
    button:hover {{
        box-shadow: 0 0 25px {GLOW_COLOR_TRANSLUCENT}, inset 0 0 10px {GLOW_COLOR_LIGHT} !important;
        transform: translateY(-2px) !important;
    }}
    
    input, select, textarea, [data-testid="stFileUploader"] {{
        border: 2px solid {GLOW_COLOR_LIGHT} !important;
        box-shadow: inset 0 0 10px {GLOW_COLOR_LIGHT}, 0 0 10px {GLOW_COLOR_LIGHT} !important;
        background-color: var(--input-background-color) !important; 
        color: var(--text-color) !important;
        transition: all 0.3s ease !important;
    }}
    
    input:focus, select:focus, textarea:focus {{
        border-color: {COLOR_PRINCIPAL} !important;
        box-shadow: inset 0 0 15px {GLOW_COLOR_LIGHT}, 0 0 20px {GLOW_COLOR_TRANSLUCENT} !important;
    }}

    .section-header {{
        background: linear-gradient(135deg, {COLOR_PRINCIPAL} 0%, {GLOW_COLOR_TRANSLUCENT} 50%, var(--background-color) 100%);
        padding: 1rem 1.5rem;
        border-radius: 10px;
        margin: 2rem 0 1rem 0;
        font-weight: 600;
        font-size: 1.3rem;
        border: 2px solid {GLOW_COLOR_TRANSLUCENT};
        box-shadow: 
            0 0 50px {GLOW_COLOR_LIGHT},
            0 0 80px {GLOW_COLOR_LIGHT},
            inset 0 0 20px rgba(255, 255, 255, 0.1);
        color: var(--text-color) !important;
    }}

    /* Winner card */
    .winner-card {{
        background: linear-gradient(135deg, rgba(102, 187, 106, 0.15) 0%, rgba(102, 187, 106, 0.05) 100%);
        border: 2px solid {COLOR_SUCCESS};
        border-radius: 10px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 5px 20px rgba(102, 187, 106, 0.3);
    }}

    /* DISTRIBUCIÓN */
    [data-testid="column"] {{
        width: 100% !important;
        padding: 0 0.5rem !important;
        box-sizing: border-box !important;
    }}

    [data-testid="stVerticalBlock"] > [data-testid="stHorizontalBlock"] {{
        gap: 1rem !important;
    }}

    /* ========== MEJORAS DEL SIDEBAR ========== */
    [data-testid="stSidebar"] {{
        background: linear-gradient(180deg, rgba(10, 14, 39, 0.95) 0%, rgba(26, 31, 58, 0.95) 50%, rgba(15, 20, 40, 0.95) 100%);
        box-shadow: -10px 0 30px rgba(0, 0, 0, 0.5), inset 2px 0 10px rgba(74, 144, 226, 0.15) !important;
        border-right: 2px solid rgba(74, 144, 226, 0.3);
    }}

    [data-testid="stSidebar"] [data-testid="stSidebarContent"] {{
        background: transparent;
        padding: 1.5rem 1rem !important;
    }}

    /* Expanders en sidebar mejorados */
    [data-testid="stSidebar"] .streamlit-expanderHeader {{
        background: linear-gradient(135deg, rgba(74, 144, 226, 0.2) 0%, rgba(90, 255, 218, 0.1) 100%);
        border: 1.5px solid rgba(74, 144, 226, 0.5);
        border-radius: 8px;
        color: #4A90E2;
        font-weight: 700;
        padding: 1rem;
        margin-bottom: 0.5rem;
        box-shadow: 0 4px 12px rgba(74, 144, 226, 0.15), inset 0 0 10px rgba(90, 255, 218, 0.08);
        transition: all 0.3s ease;
    }}

    [data-testid="stSidebar"] .streamlit-expanderHeader:hover {{
        background: linear-gradient(135deg, rgba(74, 144, 226, 0.3) 0%, rgba(90, 255, 218, 0.15) 100%);
        border-color: rgba(90, 255, 218, 0.7);
        box-shadow: 0 6px 18px rgba(74, 144, 226, 0.25), inset 0 0 15px rgba(90, 255, 218, 0.12);
    }}

    /* Contenido dentro de expanders */
    [data-testid="stSidebar"] .streamlit-expanderContent {{
        background: rgba(15, 20, 40, 0.6);
        border-radius: 8px;
        padding: 1rem !important;
        border: 1px solid rgba(90, 255, 218, 0.2);
    }}

    /* Radio buttons en sidebar */
    [data-testid="stSidebar"] [data-testid="stRadio"] {{
        padding: 0.5rem 0;
    }}

</style>
"""

st.markdown(custom_css, unsafe_allow_html=True)

# Header principal
st.markdown("""
<div class="main-header">
    <h1>EVALUACIÓN DE DISEÑOS ESP</h1>
    <p>Sistema Integral de Evaluación Técnica, Económica y Energética - Colombia</p>
</div>
""", unsafe_allow_html=True)
# Logo del sistema
st.markdown(
    "<div style='text-align:center; margin-top:8px;'>"
    "<img src='https://www.fronteraenergy.ca/wp-content/uploads/2023/05/logo-frontera-white.png' width='250'/>"
    "</div>",
    unsafe_allow_html=True
)

# ==================== PARÁMETROS ====================
SHEET_NAME = 'TLCC'
START_ROW = 9
END_ROW = 68
DATA_POINTS = END_ROW - START_ROW + 1

COLUMN_INDICES = {
    'TLCC PMM': {
        '4_5': { 'index': 14, 'title': '4 1/2"', 'key': 'TLCC PMM_4_5' },
        '3_5': { 'index': 10, 'title': '3 1/2"', 'key': 'TLCC PMM_3_5' }
    },
    'TLCC AM': {
        '4_5': { 'index': 23, 'title': '4 1/2"', 'key': 'am_4_5' },
        '3_5': { 'index': 19, 'title': '3 1/2"', 'key': 'am_3_5' }
    }
}

# ==================== FUNCIONES DE EXTRACCIÓN ====================
def extract_cost_column(df_sheet, column_index):
    if column_index < 0 or column_index >= df_sheet.shape[1]:
        vals = [0.0] * DATA_POINTS
        return { 'totalCost': 0.0, 'rawValues': vals }
    series = df_sheet.iloc[START_ROW:END_ROW+1, column_index]
    vals = pd.to_numeric(series, errors='coerce').fillna(0.0).astype(float).tolist()
    total = float(np.nansum(vals))
    return { 'totalCost': total, 'rawValues': vals }

def extract_fixed(df_sheet, column_index):
    if column_index < 0 or column_index >= df_sheet.shape[1] or df_sheet.shape[0] < 8:
        return 0.0
    val = df_sheet.iat[7, column_index]
    if pd.isna(val):
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def find_blocks_in_sheet(df_sheet):
    RANGES = [
        {"name": "Tabla 1", "range_a1": "B2:E14", "coords": [1, 13, 1, 4]},
        {"name": "Tabla 2", "range_a1": "G2:J14", "coords": [1, 13, 6, 9]},
        {"name": "Tabla 3", "range_a1": "L2:O14", "coords": [1, 13, 11, 14]},
        {"name": "Tabla 4", "range_a1": "Q2:T14", "coords": [1, 13, 16, 19]},
    ]
    blocks = []
    nrows, ncols = df_sheet.shape
    for r_def in RANGES:
        R_START, R_END, C_START, C_END = r_def["coords"]
        R_END_SAFE = min(R_END, nrows - 1)
        C_END_SAFE = min(C_END, ncols - 1)
        if R_START > R_END_SAFE or C_START > C_END_SAFE:
            continue
        data = df_sheet.iloc[R_START:R_END_SAFE+1, C_START:C_END_SAFE+1].values.tolist()
        title_cell = df_sheet.iat[R_START, C_START]
        title = f"{r_def['name']}: {str(title_cell).strip()}" if title_cell and str(title_cell).strip() else r_def['name']
        blocks.append({ 'title': title, 'range': r_def['range_a1'], 'data': data })
    return blocks

def clean_param_name(value):
    if pd.isna(value) or value is None: return ""
    s = str(value).strip().lower().replace(" ", "").replace(":", "")
    s = s.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    return s

def show_no_data_message(title="Sin Propuesta", message="No hay datos disponibles para esta sección"):
    """Muestra un mensaje elegante cuando no hay datos"""
    st.markdown(f"""
    <div class="no-data-message">
        <h3>⚠️ {title}</h3>
        <p>{message}</p>
    </div>
    """, unsafe_allow_html=True)

def process_file(uploaded_file):
    name = uploaded_file.name
    df_tlcc = None
    df_frontera = None
    diseno_data = None
    diseno_oferta = None
    diseno_oferta_meta = None

    try:
        df_tlcc = pd.read_excel(uploaded_file, sheet_name=SHEET_NAME, header=None, engine='openpyxl')
    except: pass

    try:
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet_names = wb.sheetnames
        frontera_sheet_name = next((s for s in sheet_names if s.lower().strip() == 'frontera'), None)
        diseno_sheet_name = next((s for s in sheet_names if s.lower().strip() in ['diseño', 'diseno']), None)

        if frontera_sheet_name:
            uploaded_file.seek(0)
            df_frontera = pd.read_excel(uploaded_file, sheet_name=frontera_sheet_name, header=None, engine='openpyxl')

        if diseno_sheet_name:
            uploaded_file.seek(0)
            df_diseno = pd.read_excel(uploaded_file, sheet_name=diseno_sheet_name, header=None, engine='openpyxl')
            summary_data = {}
            if df_diseno.shape[0] > 3 and df_diseno.shape[1] > 4:
                summary_data = {
                    'Proveedor': df_diseno.iat[1, 2],
                    'Pozo': df_diseno.iat[2, 2],
                    'Bomba': df_diseno.iat[3, 2],
                    'Cuerpos de Bomba': df_diseno.iat[1, 4],
                    'Numero de Etapas': df_diseno.iat[3, 4],
                }

            key_metrics = {'Mínimo': {}, 'Óptimo': {}, 'Máximo': {}}
            if df_diseno.shape[0] > 6 and df_diseno.shape[1] > 4:
                header_row = [str(cell).strip() if pd.notna(cell) else '' for cell in df_diseno.iloc[5, 1:5].values.tolist()]
                data_rows_list = df_diseno.iloc[6:40, 1:5].values.tolist()
                filtered_data = [row for row in data_rows_list if any(pd.notna(cell) and str(cell).strip() != '' for cell in row)]
                for row in filtered_data:
                    param_name = clean_param_name(row[0])
                    if not param_name: continue
                    key = None
                    # `param_name` has no spaces or accents (see clean_param_name)
                    if 'tdh' in param_name: key = 'TDH'
                    elif 'bfpd' in param_name: key = 'BFPD'
                    elif 'frecuencia' in param_name: key = 'Frecuencia'
                    elif 'kva' in param_name: key = 'kva'
                    elif 'pip' in param_name: key = 'PIP'
                    elif 'pestat' in param_name or 'pestatica' in param_name or 'p_estatica' in param_name: key = 'P estatica'
                    elif 'pdescarga' in param_name or 'p_descarga' in param_name: key = 'P descarga'
                    elif 'niveldelfluido' in param_name or 'niveldefluido' in param_name: key = 'Nivel de fluido, MD'
                    elif 'eficienciadelabomba' in param_name or 'eficienciadelbomba' in param_name or 'eficienciabomba' in param_name: key = 'Eficiencia de la bomba'
                    elif 'eficienciadelmotor' in param_name or 'eficienciamotor' in param_name: key = 'Eficiencia del motor'
                    elif 'shaftloadpump' in param_name or 'shaftload' in param_name: key = 'Shaft load Pump'
                    elif 'cargadelmotorhp' in param_name or 'cargamotorhp' in param_name or 'motorhp' in param_name: key = 'Carga del motor HP'
                    elif 'cargadelmotoramp' in param_name or 'cargamotoramp' in param_name or 'motoramp' in param_name: key = 'Carga del motor Amp'
                    elif 'cargadelazapata' in param_name or 'cargazapata' in param_name: key = 'Carga de la zapata'
                    else: continue
                    key_metrics['Mínimo'][key] = row[1] if len(row) > 1 else 'N/A'
                    key_metrics['Óptimo'][key] = row[2] if len(row) > 2 else 'N/A'
                    key_metrics['Máximo'][key] = row[3] if len(row) > 3 else 'N/A'
                diseno_data = {
                    'summary': summary_data,
                    'data': filtered_data,
                    'columns': header_row,
                    'key_metrics': key_metrics
                }

            oferta_sheet_name = next((s for s in sheet_names if s.lower().strip() in ['oferta economica','oferta_economica','ofertaeconomica','oferta']), None)
            if oferta_sheet_name:
                try:
                    sheet_of = wb[oferta_sheet_name]
                    rows = []
                    for r in range(71,81):
                        title = sheet_of.cell(row=r, column=2).value
                        valc = sheet_of.cell(row=r, column=3).value
                        vald = sheet_of.cell(row=r, column=4).value
                        if title is None and valc is None and vald is None:
                            continue
                        rows.append({'title': title if title is not None else '', 'valor': valc, 'cumple': vald})
                    if rows:
                        diseno_oferta = rows
                    try:
                        d81 = sheet_of.cell(row=81, column=4).value
                    except Exception:
                        d81 = None
                    if d81 is not None:
                        diseno_oferta_meta = {'D81': d81}
                        try:
                            if diseno_data and isinstance(diseno_data, dict):
                                diseno_data.setdefault('summary', {})
                                diseno_data['summary']['Compliance_Final_D81'] = d81
                        except Exception:
                            pass
                except Exception:
                    diseno_oferta = None
                    diseno_oferta_meta = None

            # Intentar leer hoja con valores SLA para diseño (BFPD)
            try:
                sla_sheet_name = next((s for s in sheet_names if 'data para' in s.lower() and 'dise' in s.lower() and 'sla' in s.lower()), None)
                if sla_sheet_name:
                    sheet_sla = wb[sla_sheet_name]
                    # Celdas: F23 (row 23, col 6), F18 (row 18, col 6), G23 (row 23, col 7)
                    try:
                        bfpd_min = sheet_sla.cell(row=23, column=6).value
                    except Exception:
                        bfpd_min = None
                    try:
                        bfpd_obj = sheet_sla.cell(row=18, column=6).value
                    except Exception:
                        bfpd_obj = None
                    try:
                        bfpd_max = sheet_sla.cell(row=23, column=7).value
                    except Exception:
                        bfpd_max = None

                    if diseno_oferta_meta is None:
                        diseno_oferta_meta = {}
                    diseno_oferta_meta.setdefault('SLA', {})
                    diseno_oferta_meta['SLA'].update({
                        'BFPD_min': bfpd_min,
                        'BFPD_obj': bfpd_obj,
                        'BFPD_max': bfpd_max
                    })
            except Exception:
                pass

    except Exception as e:
        st.warning(f"Error procesando hoja DISEÑO en {name}: {e}")

    result = { 'name': name, 'dataPoints': DATA_POINTS }

    if df_tlcc is not None:
        for motor_type in ['TLCC PMM', 'TLCC AM']:
            for diam_key, meta in COLUMN_INDICES[motor_type].items():
                key = meta['key']
                out = extract_cost_column(df_tlcc, meta['index'])
                fixed = extract_fixed(df_tlcc, meta['index'])
                result[key] = out
                result[f'{key}_fixed'] = fixed
    else:
        for motor_type in ['TLCC PMM', 'TLCC AM']:
            for diam_key, meta in COLUMN_INDICES[motor_type].items():
                key = meta['key']
                result[key] = { 'totalCost': 0.0, 'rawValues': [0.0]*DATA_POINTS }
                result[f'{key}_fixed'] = 0.0

    frontera_tables = []
    if df_frontera is not None:
        blocks = find_blocks_in_sheet(df_frontera)
        for b in blocks:
            pump_name = provider_name = pipe_size = "N/A"
            if len(b['data']) > 1 and len(b['data'][1]) > 1:
                provider_name = str(b['data'][1][1]).strip() if pd.notna(b['data'][1][1]) else "N/A"
            if len(b['data']) > 2 and len(b['data'][2]) > 1:
                pump_name = str(b['data'][2][1]).strip() if pd.notna(b['data'][2][1]) else "N/A"
            if len(b['data']) > 0:
                header = str(b['data'][0][0]).strip()
                if '4-1/2' in header or '4 1/2' in header or '4½' in header:
                    pipe_size = '4-1/2"'
                elif '3-1/2' in header or '3 1/2' in header or '3½' in header:
                    pipe_size = '3-1/2"'
            frontera_tables.append({
                'title': b['title'], 'range': b['range'], 'data': b['data'],
                'pumpName': pump_name, 'pipeSize': pipe_size, 'providerName': provider_name
            })
    result['fronteraTables'] = frontera_tables
    result['disenoData'] = diseno_data
    result['oferta'] = diseno_oferta
    result['oferta_meta'] = diseno_oferta_meta
    result['compliance_final_d81'] = diseno_oferta_meta.get('D81') if diseno_oferta_meta else None
    return result

# ==================== UI PRINCIPAL ====================
with st.sidebar.expander('📂 Carga de Archivos', expanded=True):
    st.markdown("""
    <div style='background: linear-gradient(135deg, rgba(74, 144, 226, 0.15) 0%, rgba(74, 144, 226, 0.05) 100%);
                padding: 1rem; border-radius: 10px; margin-bottom: 1rem;
                border: 1.5px solid rgba(74, 144, 226, 0.4);
                box-shadow: 0 4px 12px rgba(74, 144, 226, 0.1);'>
        <p style='color: #4A90E2; font-weight: 600; margin: 0 0 0.5rem 0; font-size: 0.95rem;'>📥 Propuestas Excel</p>
        <p style='margin:0; font-size:0.85rem; color: rgba(255,255,255,0.7);'>Arrastra o selecciona (.xlsx, .xls)</p>
    </div>
    """, unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        'Selecciona propuestas Excel', 
        type=['xlsx', 'xls'], 
        accept_multiple_files=True,
        key='uploaded_files',
        help="Sube para comparar ofertas"
    )

if not uploaded_files:
    st.info('📥 Sube uno o más archivos Excel para iniciar la evaluación.')
    st.stop()

for f in uploaded_files:
    f.seek(0)

with st.spinner('⏳ Procesando archivos...'):
    processed = []
    for f in uploaded_files:
        try:
            processed.append(process_file(f))
        except Exception as e:
            st.error(f'❌ Error procesando {f.name}: {e}')

if len(processed) == 0:
    st.error('❌ No se extrajeron datos válidos.')
    st.stop()

menu_options = [
    ('📋 Resumen ', 'RESUMEN EJECUTIVO'),
    ('⚡ TLCC PMM', 'TLCC PMM'),
    ('⚡ TLCC AM', 'TLCC AM'),
    ('💰 Costos', 'COSTOS'),
    ('🛠️ EV Técnica', 'EV TECNICA')
]

with st.sidebar.expander('📚 Menú de Análisis', expanded=True):
    st.markdown("""
    <div style='background: linear-gradient(135deg, rgba(90, 255, 218, 0.15) 0%, rgba(90, 255, 218, 0.05) 100%);
                padding: 1rem; border-radius: 10px; margin-bottom: 1rem;
                border: 1.5px solid rgba(90, 255, 218, 0.4);
                box-shadow: 0 4px 12px rgba(90, 255, 218, 0.1);'>
        <p style='color: #5AFFDA; font-weight: 600; margin: 0; font-size: 0.95rem;'>Selecciona el tipo de evaluación</p>
    </div>
    """, unsafe_allow_html=True)
    menu_labels = [m[0] for m in menu_options]
    sel_label = st.radio('Análisis', menu_labels, index=0, help='Selecciona el tipo de análisis a realizar', label_visibility="collapsed")

# motor_choice mantiene la clave interna usada en el resto del código
motor_choice = next(key for label, key in menu_options if label == sel_label)
motor_label = sel_label

st.markdown('<div class="section-header">📊 Resumen de Archivos Procesados</div>', unsafe_allow_html=True)
cols = st.columns(3)
cols[0].metric('📁 Archivos', len(processed), help="Total de archivos cargados")
cols[1].metric('📈 Puntos de Datos', DATA_POINTS, help="Puntos de análisis por archivo")
cols[2].metric('⚙️ Tipo de Análisis', motor_choice, help="Configuración seleccionada")

# ==================== SECCIÓN RESUMEN EJECUTIVO ====================
if motor_choice == 'RESUMEN EJECUTIVO':
   
    
    # Recolectar todos los datos
    ROW_MAP = {
        'NOMBRE BOMBA': 2,
        'EQUIPO DE SUPERFICIE': 3,
        'EQUIPO ESP': 4,
        'Cable + Accesorios': 5,
        'Y-TOOL': 6,
        'Servicios': 8,
        'EQUIPO + SERVICIO': 9,
        'TOTAL': 10,
        'TLCC': 11,
        'EQUIPO+SERVICIOS+TLCC': 12
    }

    # Extraer datos económicos
    full_offers = []
    rrigo_offers = []
    alt_offers = []

    for p in processed:
        for t in p.get('fronteraTables', []):
            data = t['data']
            if not data or len(data) < 13: continue

            provider = t['providerName']
            pump = t['pumpName']
            pipe = t['pipeSize']
            title = f"{provider} - {pump}"

            total_row = None
            for row in data:
                if len(row) >= 1 and "equipo+servicios+tlcc" in str(row[0]).lower():
                    total_row = row
                    break
            if not total_row or len(total_row) < 4: continue

            def clean_val(val):
                if pd.isna(val): return 0.0
                s = str(val).replace(", ", "").replace(",", "").replace("USD", "").strip()
                try: return float(s)
                except: return 0.0

            full = clean_val(total_row[1])
            rrigo = clean_val(total_row[2])
            alt = clean_val(total_row[3])

            offer = {
                'title': title,
                'provider': provider,
                'pump': pump,
                'pipe': pipe,
                'filename': p['name']
            }

            if full > 0:
                full_offers.append({**offer, 'value': full})
            if rrigo > 0:
                rrigo_offers.append({**offer, 'value': rrigo})
            if alt > 0:
                alt_offers.append({**offer, 'value': alt})

    # Extraer datos técnicos
    technical_data = []
    for p in processed:
        diseno = p.get('disenoData') or {}
        summary = diseno.get('summary', {})
        km = diseno.get('key_metrics', {})
        
        compliance_final_d81 = p.get('compliance_final_d81')
        if compliance_final_d81 is None:
            compliance_final_d81 = (p.get('oferta_meta') or {}).get('D81')
        if compliance_final_d81 is None:
            compliance_final_d81 = summary.get('Compliance_Final_D81')
        
        def _is_cumple(val):
            """Determina si 'val' indica que se cumple la condición.
            - Si contiene un porcentaje numérico (ej. '85% cumplimiento'), se compara contra 80% (mayor que 80 => cumple).
            - Si contiene palabras explícitas ('cumple','si','sí','ok', etc.) devuelve True.
            - Valores numéricos se tratan como booleanos (no cero => True).
            """
            import re
            if val is None:
                return False
            if isinstance(val, (int, float)):
                return bool(val)
            s = str(val).strip().lower()
            # Detectar porcentaje numérico
            m = re.search(r"(\d+(?:[\.,]\d+)?)\s*%", s)
            if m:
                try:
                    num = float(m.group(1).replace(',', '.'))
                    return num > 80.0
                except Exception:
                    pass
            # Palabras que indican cumplimiento
            positives = ['si', 'sí', 'cumple', 'ok', 'yes', 'true', 'cumplido']
            negatives = ['no cumple', 'no cumple', 'no', 'false', 'falso']
            # Si hay una negación explícita, consideramos no cumple
            for neg in negatives:
                if neg in s:
                    return False
            return any(x in s for x in positives)
        
        technical_data.append({
            'Archivo': p['name'],
            'Proveedor': summary.get('Proveedor', 'N/A'),
            'Bomba': summary.get('Bomba', 'N/A'),
            'Cumplimiento': '✅ CUMPLE' if _is_cumple(compliance_final_d81) else '❌ NO CUMPLE',
            'TDH_Optimo': km.get('Óptimo', {}).get('TDH', 'N/A'),
            'BFPD_Optimo': km.get('Óptimo', {}).get('BFPD', 'N/A'),
            'Frecuencia_Optimo': km.get('Óptimo', {}).get('Frecuencia', 'N/A')
        })

    # Extraer datos TLCC
    tlcc_pmm_data = []
    tlcc_am_data = []
    
    for p in processed:
        tlcc_pmm_data.append({
            'Proveedor': p['name'],
            '4½" Total': p['TLCC PMM_4_5']['totalCost'],
            '4½" Fijo': p['TLCC PMM_4_5_fixed'],
            '3½" Total': p['TLCC PMM_3_5']['totalCost'],
            '3½" Fijo': p['TLCC PMM_3_5_fixed']
        })
        
        tlcc_am_data.append({
            'Proveedor': p['name'],
            '4½" Total': p['am_4_5']['totalCost'],
            '4½" Fijo': p['am_4_5_fixed'],
            '3½" Total': p['am_3_5']['totalCost'],
            '3½" Fijo': p['am_3_5_fixed']
        })

    # ===== SISTEMA DE SCORING Y RANKING (refactorizado para AM/PMM) =====

    allowed_tables = {
        'AM': [3, 4],
        'PMM': [2, 1]
    }

    def compute_scores(view):
        # view: 'AM' or 'PMM'
        provider_scores = {}
        for p in processed:
            provider_scores[p['name']] = {
                'economico': 0,
                'tecnico': 0,
                'energetico': 0,
                'total': 0,
                'details': {}
            }

        # 1) SCORING ECONÓMICO (40 pts) - usar solo tablas permitidas para la vista
        all_economic_offers = []
        for p in processed:
            for t in p.get('fronteraTables', []):
                data = t['data']
                if not data or len(data) < 1:
                    continue
                title_lower = (t.get('title') or '').lower()
                m = re.search(r'tabla\s*(\d+)', title_lower)
                table_num = int(m.group(1)) if m else None
                if table_num is None or table_num not in allowed_tables.get(view, []):
                    continue

                provider = t.get('providerName')
                pump = t.get('pumpName')
                pipe = t.get('pipeSize')

                total_row = None
                for row in data:
                    if len(row) >= 1 and 'equipo+servicios+tlcc' in str(row[0]).lower():
                        total_row = row
                        break
                if not total_row or len(total_row) < 4:
                    continue

                def clean_val(v):
                    if pd.isna(v): return 0.0
                    s = str(v).replace('$', '').replace(',', '').replace('USD', '').strip()
                    try: return float(s)
                    except: return 0.0

                for i, cat in enumerate(['FULL PRICE', 'R-R-I-G-O', 'ALT AHORRO'], 1):
                    val = clean_val(total_row[i])
                    if val > 0:
                        all_economic_offers.append({
                            'provider': provider,
                            'pump': pump,
                            'pipe': pipe,
                            'category': cat,
                            'value': val,
                            'filename': p['name']
                        })

        # calcular PEF (Puntuación Económica Final) basado en las 3 categorías
        if all_economic_offers:
            df_econ = pd.DataFrame(all_economic_offers)
            categories_list = ['FULL PRICE', 'R-R-I-G-O', 'ALT AHORRO']
            # categorías válidas (las que tienen al menos una oferta)
            valid_categories = [c for c in categories_list if c in df_econ['category'].unique()]
            num_valid = len(valid_categories)

            # puntos por posición dentro de cada categoría
            points_map = [40, 35, 30, 25, 20, 15, 10, 5]

            # preparar estructura para puntos por categoría
            per_cat_points = {c: {} for c in valid_categories}
            providers = [p['name'] for p in processed if p.get('name')]

            # asignar puntos por posición en cada categoría (mejor valor -> 1er puesto)
            for cat in valid_categories:
                df_cat = df_econ[df_econ['category'] == cat]
                if df_cat.empty:
                    continue
                # por proveedor tomar la mejor (mínima) oferta en la categoría
                df_best = df_cat.groupby('filename', as_index=False)['value'].min()
                df_sorted = df_best.sort_values('value').reset_index(drop=True)
                for idx, row in df_sorted.iterrows():
                    pos = idx
                    pts = points_map[pos] if pos < len(points_map) else 0
                    per_cat_points[cat][row['filename']] = pts

            # construir PBO (Puntaje Bruto Obtenido) por proveedor sumando por categorías válidas
            pbo_map = {prov: 0.0 for prov in providers}
            # si un proveedor no tiene oferta en una categoría válida, recibe 10 pts para esa categoría
            for cat in valid_categories:
                for prov in providers:
                    pts = per_cat_points.get(cat, {}).get(prov)
                    if pts is None:
                        # si el proveedor no participó en la categoría, 10 puntos mínimos
                        pts = 10
                    pbo_map[prov] = pbo_map.get(prov, 0.0) + float(pts)

            # PBP = 40 * número de categorías válidas
            pbp = 40 * num_valid if num_valid > 0 else 0

            # calcular PD (porcentaje de desempeño) por proveedor
            pd_map = {}
            for prov, pbo in pbo_map.items():
                pd_pct = (pbo / pbp * 100.0) if pbp > 0 else 0.0
                pd_map[prov] = pd_pct

            # Asignar puntos económicos discretos por posición basada en PBO
            # Ordenar proveedores por PBO (desc)
            sorted_by_pbo = sorted(pbo_map.items(), key=lambda x: (x[1], x[0]), reverse=True)
            rank_points = [40, 35, 30, 25, 20, 15, 10, 5]
            assigned_points = {}
            for idx, (prov, pbo_val) in enumerate(sorted_by_pbo):
                pts = rank_points[idx] if idx < len(rank_points) else 0
                assigned_points[prov] = pts

            # Guardar en provider_scores
            for prov in pbo_map.keys():
                provider_scores.setdefault(prov, {
                    'economico': 0,
                    'tecnico': 0,
                    'energetico': 0,
                    'total': 0,
                    'details': {}
                })
                provider_scores[prov]['economico'] = float(assigned_points.get(prov, 0))
                provider_scores[prov]['details']['PBO'] = float(pbo_map.get(prov, 0.0))
                provider_scores[prov]['details']['PD'] = float(pd_map.get(prov, 0.0))
                provider_scores[prov]['details']['Rank_Econ'] = int([i for i, (pp, _) in enumerate(sorted_by_pbo) if pp == prov][0]) + 1
                provider_scores[prov]['details']['Pts_Assigned'] = float(assigned_points.get(prov, 0))

            # Mostrar tabla de diagnóstico económica para depuración (opcional)
            try:
                diag_rows = []
                for prov in providers:
                    row = {'Proveedor': prov, 'PBO': pbo_map.get(prov, 0.0), 'PD': pd_map.get(prov, 0.0), 'PEF': pef_map.get(prov, 0.0)}
                    # añadir puntos por categoría
                    for cat in valid_categories:
                        row[f'Pts_{cat}'] = per_cat_points.get(cat, {}).get(prov, 10)
                    diag_rows.append(row)
                df_diag = pd.DataFrame(diag_rows)
                st.expander('📋 Diagnóstico Económico (PBO / PD / PEF) - desplegar para ver detalles', expanded=False)
                with st.expander('📋 Diagnóstico Económico (PBO / PD / PEF) - desplegar para ver detalles'):
                    st.dataframe(df_diag.sort_values('PEF', ascending=False).reset_index(drop=True))
            except Exception:
                pass

        # 2) SCORING TÉCNICO (35 pts) - igual para AM y PMM
        for p in processed:
            diseno = p.get('disenoData') or {}
            summary = diseno.get('summary', {})
            compliance_final_d81 = p.get('compliance_final_d81')
            if compliance_final_d81 is None:
                compliance_final_d81 = (p.get('oferta_meta') or {}).get('D81')
            if compliance_final_d81 is None:
                compliance_final_d81 = summary.get('Compliance_Final_D81')

            def _is_cumple(val):
                if val is None:
                    return False
                if isinstance(val, (int, float)):
                    return bool(val)
                s = str(val).strip().lower()
                m = re.search(r"(\d+(?:[\.,]\d+)?)\s*%", s)
                if m:
                    try:
                        num = float(m.group(1).replace(',', '.'))
                        return num > 80.0
                    except Exception:
                        pass
                positives = ['si', 'sí', 'cumple', 'ok', 'yes', 'true', 'cumplido']
                negatives = ['no cumple', 'no', 'false', 'falso']
                for neg in negatives:
                    if neg in s:
                        return False
                return any(x in s for x in positives)

            if _is_cumple(compliance_final_d81):
                provider_scores[p['name']]['tecnico'] = 35
                provider_scores[p['name']]['details']['tech_status'] = 'CUMPLE'
            else:
                provider_scores[p['name']]['tecnico'] = 0
                provider_scores[p['name']]['details']['tech_status'] = 'NO CUMPLE'

        # 3) SCORING ENERGÉTICO (25 pts) - usar solo AM o solo PMM según view
        all_energy_costs = []
        for p in processed:
            if view == 'AM':
                v45 = p.get('am_4_5', {}).get('totalCost', 0)
                v35 = p.get('am_3_5', {}).get('totalCost', 0)
            else:
                v45 = p.get('TLCC PMM_4_5', {}).get('totalCost', 0)
                v35 = p.get('TLCC PMM_3_5', {}).get('totalCost', 0)
            avg_cost = np.mean([x for x in [v45, v35] if x > 0]) if any(x > 0 for x in [v45, v35]) else 0
            if avg_cost > 0:
                all_energy_costs.append({'provider': p['name'], 'avg_cost': avg_cost})

        if all_energy_costs:
            df_energy = pd.DataFrame(all_energy_costs).sort_values('avg_cost').reset_index(drop=True)
            energy_points = [25, 20, 15, 10, 5, 0]
            values = df_energy['avg_cost'].tolist()
            first_pos = {}
            for i, v in enumerate(values):
                if v not in first_pos:
                    first_pos[v] = i
            for idx, row in df_energy.iterrows():
                rank_idx = first_pos[row['avg_cost']]
                pts = energy_points[rank_idx] if rank_idx < len(energy_points) else 0
                provider_scores[row['provider']]['energetico'] = pts
                provider_scores[row['provider']]['details']['energy_avg'] = row['avg_cost']

        # totales y orden
        for prov in provider_scores:
            provider_scores[prov]['total'] = (
                provider_scores[prov]['economico'] +
                provider_scores[prov]['tecnico'] +
                provider_scores[prov]['energetico']
            )

        ranked = sorted(provider_scores.items(), key=lambda x: x[1]['total'], reverse=True)
        return provider_scores, ranked

    # UI: botones para evaluar AM y PMM por separado
    st.markdown('### Evaluar por tipo de motor')
    eval_cols = st.columns(2)
    am_pressed = eval_cols[0].button('Evaluar AM')
    pmm_pressed = eval_cols[1].button('Evaluar PMM')

    # Si se presiona un botón, guardamos la vista en session_state y la usamos
    if am_pressed:
        prov_scores_am, ranked_am = compute_scores('AM')
        ranked_providers = ranked_am
        provider_scores = prov_scores_am
        st.session_state['evaluation_view'] = 'AM'
        current_view = 'AM'
        st.success('✅ Evaluación AM completada')
    elif pmm_pressed:
        prov_scores_pmm, ranked_pmm = compute_scores('PMM')
        ranked_providers = ranked_pmm
        provider_scores = prov_scores_pmm
        st.session_state['evaluation_view'] = 'PMM'
        current_view = 'PMM'
        st.success('✅ Evaluación PMM completada')
    else:
        # usar la vista almacenada en session_state o PMM por defecto
        current_view = st.session_state.get('evaluation_view', 'PMM')
        provider_scores, ranked_providers = compute_scores(current_view)
    
    # ===== CARD DEL GANADOR =====
    if ranked_providers:
        winner_name, winner_data = ranked_providers[0]
        
        st.markdown(f"""
        <div style='background: linear-gradient(155deg, #08004D 0%, #00378A 100%);
                    padding: 2rem; border-radius: 25px; margin: 2rem 0;
                    box-shadow: 0 10px 50px rgba(0, 55, 138, 0.4);
                    border: 3px solid #00378A;'>
            <div style='text-align: center;'>
                <h2 style='color: white; font-size: 2rem; margin-bottom: 1rem;'>
                    🏆 PROVEEDOR RECOMENDADO
                </h2>
                <h1 style='color: white; font-size: 3rem; margin: 1rem 0; font-weight: 800;'>
                    {winner_name}
                </h1>
                <div style='display: flex; justify-content: center; gap: 2rem; margin-top: 2rem;'>
                    <div style='background: rgba(255,255,255,0.2); padding: 1rem 2rem; border-radius: 10px;'>
                        <p style='color: rgba(255,255,255,0.9); margin: 0; font-size: 0.9rem;'>PUNTAJE TOTAL</p>
                        <h2 style='color: white; margin: 0.5rem 0 0 0; font-size: 2.5rem;'>{winner_data['total']:.1f}</h2>
                        <p style='color: rgba(255,255,255,0.8); margin: 0; font-size: 0.8rem;'>de 100 puntos</p>
                    </div>
                    <div style='background: rgba(255,255,255,0.2); padding: 1rem 2rem; border-radius: 10px;'>
                        <p style='color: rgba(255,255,255,0.9); margin: 0; font-size: 0.9rem;'>ECONÓMICO</p>
                        <h2 style='color: white; margin: 0.5rem 0 0 0; font-size: 2.5rem;'>{winner_data['economico']:.1f}</h2>
                    </div>
                    <div style='background: rgba(255,255,255,0.2); padding: 1rem 2rem; border-radius: 10px;'>
                        <p style='color: rgba(255,255,255,0.9); margin: 0; font-size: 0.9rem;'>TÉCNICO</p>
                        <h2 style='color: white; margin: 0.5rem 0 0 0; font-size: 2.5rem;'>{winner_data['tecnico']:.1f}</h2>
                    </div>
                    <div style='background: rgba(255,255,255,0.2); padding: 1rem 2rem; border-radius: 10px;'>
                        <p style='color: rgba(255,255,255,0.9); margin: 0; font-size: 0.9rem;'>ENERGÉTICO</p>
                        <h2 style='color: white; margin: 0.5rem 0 0 0; font-size: 2.5rem;'>{winner_data['energetico']:.1f}</h2>
                    </div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # ===== RANKING COMPLETO CON CARDS =====
    st.markdown("### 📊 RANKING COMPLETO DE PROVEEDORES")
    
    for idx, (prov_name, prov_data) in enumerate(ranked_providers, 1):
        # Definir color según posición
        if idx == 1:
            color_start, color_end = "#057831", "#02C222"
            medal = "🥇"
        elif idx == 2:
            color_start, color_end = "#00378A", "#8e54e9"
            medal = "🥈"
        elif idx == 3:
            color_start, color_end = "#f857a6", "#FC4300"
            medal = "🥉"
        else:
            color_start, color_end = "#52520D", "#E0E000"
            medal = f"#{idx}"
        
        tech_status = prov_data['details'].get('tech_status', 'N/A')
        tech_icon = "✅" if tech_status == "CUMPLE" else "❌"
        
        # Card compacta y elegante
        col1, col2 = st.columns([3, 1])
        
        with col1:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, {color_start} 0%, {color_end} 100%);
                        padding: 1.5rem; border-radius: 12px; margin-bottom: 1rem;
                        box-shadow: 0 5px 20px rgba(0,0,0,0.2);'>
                <div style='display: flex; justify-content: space-between; align-items: center;'>
                    <div>
                        <h3 style='color: white; margin: 0; font-size: 1.8rem;'>
                            {medal} {prov_name}
                        </h3>
                        <p style='color: rgba(255,255,255,0.9); margin: 0.5rem 0 0 0;'>
                            Cumplimiento Técnico: {tech_icon} {tech_status}
                        </p>
                    </div>
                    <div style='text-align: right;'>
                        <p style='color: rgba(255,255,255,0.8); margin: 0; font-size: 0.9rem;'>PUNTAJE TOTAL</p>
                        <h1 style='color: white; margin: 0.3rem 0 0 0; font-size: 3rem; font-weight: 800;'>
                            {prov_data['total']:.1f}
                        </h1>
                    </div>
                </div>
                <div style='display: flex; gap: 1rem; margin-top: 1rem;'>
                    <div style='flex: 1; background: rgba(255,255,255,0.15); padding: 0.8rem; border-radius: 8px;'>
                        <p style='color: rgba(255,255,255,0.9); margin: 0; font-size: 0.85rem;'>💰 Económico</p>
                        <h3 style='color: white; margin: 0.3rem 0 0 0;'>{prov_data['economico']:.1f}/40</h3>
                    </div>
                    <div style='flex: 1; background: rgba(255,255,255,0.15); padding: 0.8rem; border-radius: 8px;'>
                        <p style='color: rgba(255,255,255,0.9); margin: 0; font-size: 0.85rem;'>🔧 Técnico</p>
                        <h3 style='color: white; margin: 0.3rem 0 0 0;'>{prov_data['tecnico']:.1f}/35</h3>
                    </div>
                    <div style='flex: 1; background: rgba(255,255,255,0.15); padding: 0.8rem; border-radius: 8px;'>
                        <p style='color: rgba(255,255,255,0.9); margin: 0; font-size: 0.85rem;'>⚡ Energético</p>
                        <h3 style='color: white; margin: 0.3rem 0 0 0;'>{prov_data['energetico']:.1f}/25</h3>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            # Gráfico de radar minimalista con el puntaje
            fig = go.Figure()
            
            categories = ['Económico<br>(40pts)', 'Técnico<br>(35pts)', 'Energético<br>(25pts)']
            values = [
                prov_data['economico']/40*100,
                prov_data['tecnico']/35*100,
                prov_data['energetico']/25*100
            ]
            
            fig.add_trace(go.Scatterpolar(
                r=values + [values[0]],
                theta=categories + [categories[0]],
                fill='toself',
                fillcolor=f'rgba({int(color_start[1:3], 16)}, {int(color_start[3:5], 16)}, {int(color_start[5:7], 16)}, 0.4)',
                line=dict(color=color_start, width=3),
                name=prov_name
            ))
            
            fig.update_layout(
                polar=dict(
                    radialaxis=dict(
                        visible=True,
                        range=[0, 100],
                        showticklabels=False,
                        gridcolor='rgba(255,255,255,0.2)'
                    ),
                    angularaxis=dict(
                        gridcolor='rgba(255,255,255,0.2)'
                    ),
                    bgcolor='rgba(0,0,0,0)'
                ),
                showlegend=False,
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=40, r=40, t=40, b=40),
                height=220,
                font=dict(color='white', size=9)
            )
            
            st.plotly_chart(fig, use_container_width=True)
    
    st.markdown("---")
    
    # ===== CLASIFICACIÓN POR ALTERNATIVAS ECONÓMICAS (3 RANKINGS) =====
    st.markdown('<div class="section-header">🏷️ Clasificación por Alternativa Económica</div>', unsafe_allow_html=True)
    # Reconstruir `all_economic_offers` a partir de `processed`, filtrando por la vista seleccionada
    if 'all_economic_offers' not in locals() or not all_economic_offers:
        all_economic_offers = []
        view_sel = current_view if 'current_view' in locals() else st.session_state.get('evaluation_view', 'PMM')
        for p in processed:
            for t in p.get('fronteraTables', []):
                data = t.get('data')
                if not data or len(data) < 1:
                    continue
                title_lower = (t.get('title') or '').lower()
                m = re.search(r'tabla\s*(\d+)', title_lower)
                table_num = int(m.group(1)) if m else None
                if table_num is None or table_num not in allowed_tables.get(view_sel, []):
                    continue

                provider = t.get('providerName')
                pump = t.get('pumpName')
                pipe = t.get('pipeSize')
                total_row = None
                for row in data:
                    if len(row) >= 1 and "equipo+servicios+tlcc" in str(row[0]).lower():
                        total_row = row
                        break
                if not total_row or len(total_row) < 4:
                    continue

                def clean_val(val):
                    if pd.isna(val): return 0.0
                    s = str(val).replace("$", "").replace(", ", "").replace(",", "").replace("USD", "").strip()
                    try: return float(s)
                    except: return 0.0

                for i, cat in enumerate(['FULL PRICE', 'R-R-I-G-O', 'ALT AHORRO'], 1):
                    val = clean_val(total_row[i])
                    if val > 0:
                        all_economic_offers.append({
                            'provider': provider,
                            'pump': pump,
                            'pipe': pipe,
                            'category': 'FULL PRICE' if i==1 else ('R-R-I-G-O' if i==2 else 'ALT AHORRO'),
                            'value': val,
                            'filename': p['name']
                        })

    def compute_category_ranking(category, all_offers, base_provider_scores=None):
        """Ranking económico por categoría.
        - Solo considera la parte económica (sin técnico/energético).
        - Para cada tubería (pipe) asigna puntos fijos por posición: 1->40,2->35,3->30,4->25,...
        - Suma puntos por proveedor a través de las tuberías y devuelve lista ordenada.
        """
        points_map = [40, 35, 30, 25, 20, 15, 10, 5]
        df_cat = pd.DataFrame([o for o in all_offers if o['category'] == category])
        if df_cat.empty:
            return []

        econ_scores = {}
        # Para cada tubería, ordenar por valor ascendente y asignar puntos por posición
        for pipe, df_group in df_cat.groupby('pipe'):
            df_sorted = df_group.sort_values('value').reset_index(drop=True)
            for idx, row in df_sorted.iterrows():
                pos = idx  # 0-based
                pts = points_map[pos] if pos < len(points_map) else 0
                fname = row['filename']
                econ_scores[fname] = econ_scores.get(fname, 0) + pts
        # Asegurar que todos los proveedores aparezcan en el ranking.
        all_providers = [p.get('name') for p in processed if p.get('name')]
        # proveedores que sí tuvieron oferta en esta categoría
        providers_with_offer = set(econ_scores.keys())

        # Construir ranking: si no hay oferta, asignar 10 puntos y marcar has_offer=False
        ranking = []
        for prov in all_providers:
            if prov in providers_with_offer:
                pts = econ_scores.get(prov, 0)
                has_offer = True
            else:
                pts = 10
                has_offer = False
            ranking.append({'provider': prov, 'econ_points': pts, 'has_offer': has_offer})

        ranking = sorted(ranking, key=lambda x: x['econ_points'], reverse=True)
        return ranking

    # Crear tres rankings: FULL PRICE, R-R-I-G-O, ALTERNATIVA AHORRO
    base_scores = provider_scores  # ya contiene tecnico y energetico calculados antes
    rp_full = compute_category_ranking('FULL PRICE', all_economic_offers, base_scores)
    rp_rrigo = compute_category_ranking('R-R-I-G-O', all_economic_offers, base_scores)
    rp_alt = compute_category_ranking('ALT AHORRO', all_economic_offers, base_scores)

    cols_rank = st.columns(3)
    cat_info = [
        ('💎 FULL PRICE', rp_full, '#FC4300', '#FF9CF2'),
        ('🔄 R-R-I-G-O', rp_rrigo, '#057831', '#02C222'),
        ('💰 ALTERNATIVA AHORRO', rp_alt, '#00378A', '#5C8EFF')
    ]
    for i, (label, ranking_list, cstart, cend) in enumerate(cat_info):
        with cols_rank[i]:
            st.markdown(f"**{label}**")
            if not ranking_list:
                show_no_data_message('Sin Datos', 'No hay propuestas')
                continue
            # Mostrar top 5 con tarjetas compactas (solo económico)
            for j, item in enumerate(ranking_list[:5], 1):
                prov = item.get('provider')
                econ_pts = item.get('econ_points', 0)
                has_offer = item.get('has_offer', True)
                nota = "" if has_offer else " — sin oferta"
                display_sub = f"Puntos Econ: {econ_pts}"
                if not has_offer:
                    display_sub += " (sin oferta)"
                st.markdown(
                    f"<div style='background: linear-gradient(135deg, {cstart} 0%, {cend} 100%); padding:10px; border-radius:10px; margin-bottom:8px;'>"
                    f"<div style='display:flex; justify-content:space-between; align-items:center;'>"
                    f"<div><strong style='font-size:1.05rem;'>{j}. {prov}{nota}</strong><div style='font-size:0.85rem; opacity:0.9;'>{display_sub}</div></div>"
                    f"<div style='text-align:right;'><div style='font-weight:800; font-size:1.2rem;'>{econ_pts}</div></div>"
                    f"</div></div>", unsafe_allow_html=True)

    st.markdown('---')
    
    # ===== ANÁLISIS DETALLADO POR CRITERIO =====
    tab1, tab2, tab3 = st.tabs(["💰 Análisis Económico", "🔧 Análisis Técnico", "⚡ Análisis Energético"])
    
    with tab1:
        st.markdown("#### Comparativa de Ofertas Económicas")
        # Reemplazamos la comparativa por los mismos datos que se muestran
        # en la sección `COSTOS` -> "🏆 Ganadores por Categoría y Tubería".
        # Construir listas por categoría (misma lógica que en COSTOS)
        full_offers = []
        rrigo_offers = []
        alt_offers = []

        view_sel = current_view if 'current_view' in locals() else st.session_state.get('evaluation_view', 'PMM')
        for p in processed:
            for t in p.get('fronteraTables', []):
                data = t.get('data')
                if not data or len(data) < 1:
                    continue

                # Filtrar tablas según la vista seleccionada (AM/PMM)
                title_lower = (t.get('title') or '').lower()
                m = re.search(r'tabla\s*(\d+)', title_lower)
                table_num = int(m.group(1)) if m else None
                if table_num is None or table_num not in allowed_tables.get(view_sel, []):
                    continue

                provider = t.get('providerName')
                pump = t.get('pumpName')
                pipe = t.get('pipeSize')
                title = f"{provider} - {pump}"

                total_row = None
                for row in data:
                    if len(row) >= 1 and "equipo+servicios+tlcc" in str(row[0]).lower():
                        total_row = row
                        break
                if not total_row or len(total_row) < 4:
                    continue

                def clean_val(val):
                    if pd.isna(val): return 0.0
                    s = str(val).replace(", ", "").replace(",", "").replace("USD", "").strip()
                    try: return float(s)
                    except: return 0.0

                full = clean_val(total_row[1])
                rrigo = clean_val(total_row[2])
                alt = clean_val(total_row[3])

                offer = {
                    'title': title,
                    'provider': provider,
                    'pump': pump,
                    'pipe': pipe,
                    'filename': p['name']
                }

                if full > 0:
                    full_offers.append({**offer, 'value': full})
                if rrigo > 0:
                    rrigo_offers.append({**offer, 'value': rrigo})
                if alt > 0:
                    alt_offers.append({**offer, 'value': alt})

        # Mostrar los ganadores por categoría y tubería (misma presentación que COSTOS)
        # Envolver en una "cajita" visual y mostrar ambas tuberías (3½" y 4½")
        st.markdown('<div class="winner-card" style="padding:1rem;">', unsafe_allow_html=True)
        cols = st.columns(3)
        winners = {}

        def _get_best_by_pipe_from_list(offers_list, pipe_size):
            try:
                df_tmp = pd.DataFrame(offers_list)
            except Exception:
                return None
            if df_tmp.empty: return None
            df_pipe = df_tmp[df_tmp['pipe'] == pipe_size]
            if df_pipe.empty: return None
            return df_pipe.sort_values('value').iloc[0]

        for i, (cat_name, offers) in enumerate([
            ('💎 FULL PRICE', full_offers),
            ('🔄 R-R-I-G-O', rrigo_offers),
            ('💰 ALTERNATIVA AHORRO', alt_offers)
        ]):
            with cols[i]:
                st.markdown(f"**{cat_name}**")
                b35 = _get_best_by_pipe_from_list(offers, '3-1/2"')
                b45 = _get_best_by_pipe_from_list(offers, '4-1/2"')

                # Mostrar dos sub-cajas (3½" y 4½") lado a lado con estilo dinámico
                left, right = st.columns([1,1])
                with left:
                    if b35 is not None:
                        prov = b35['provider']
                        pump = b35.get('pump', 'N/A')
                        val = b35['value']
                        st.markdown(
                            f"<div style='background: linear-gradient(135deg, #667eea 0%, #8a7df6 100%); padding:12px; border-radius:10px; color:white;'>"
                            f"<h4 style='margin:0;'>3½ \" — <small style=\"opacity:0.95;\">{prov}</small></h4>"
                            f"<p style='margin:6px 0 0 0; font-size:1.1rem; font-weight:700;'>${val:,.0f}</p>"
                            f"<p style='margin:6px 0 0 0; font-size:0.85rem; opacity:0.9;'>Bomba: {pump}</p>"
                            f"</div>", unsafe_allow_html=True)
                    else:
                        show_no_data_message('Sin Propuesta 3½"', 'No hay ofertas')

                with right:
                    if b45 is not None:
                        prov = b45['provider']
                        pump = b45.get('pump', 'N/A')
                        val = b45['value']
                        st.markdown(
                            f"<div style='background: linear-gradient(135deg, #43e97b 0%, #38ef7d 100%); padding:12px; border-radius:20px; color:white;'>"
                            f"<h4 style='margin:0;'>4½ \" — <small style=\"opacity:0.95;\">{prov}</small></h4>"
                            f"<p style='margin:6px 0 0 0; font-size:1.1rem; font-weight:700;'>${val:,.0f}</p>"
                            f"<p style='margin:4px 0 0 0; font-size:0.85rem; opacity:0.9;'>Bomba: {pump}</p>"
                            f"</div>", unsafe_allow_html=True)
                    else:
                        show_no_data_message('Sin Propuesta 4½"', 'No hay ofertas')

        st.markdown('---')
        
        # Además mostrar una tabla compacta con los ganadores si existen
        rows = []
        for cat, pair in winners.items():
            b35, b45 = pair
            if b35 is not None:
                rows.append({'Categoría': cat, 'Tubería': '3-1/2"', 'Proveedor - Bomba': b35['title'], 'Valor (USD)': b35['value']})
            if b45 is not None:
                rows.append({'Categoría': cat, 'Tubería': '4-1/2"', 'Proveedor - Bomba': b45['title'], 'Valor (USD)': b45['value']})

        if rows:
            df_winners = pd.DataFrame(rows).sort_values(['Categoría','Tubería'])
            df_winners['Valor (USD)'] = df_winners['Valor (USD)'].apply(lambda x: f"${x:,.0f}")
            st.dataframe(df_winners, use_container_width=True, hide_index=True)
    
    with tab2:
        st.markdown("#### Estado de Cumplimiento Técnico")
        
        technical_summary = []
        for p in processed:
            diseno = p.get('disenoData') or {}
            summary = diseno.get('summary', {})
            compliance = p.get('compliance_final_d81')
            
            def _is_cumple(val):
                import re
                if val is None:
                    return False
                if isinstance(val, (int, float)):
                    return bool(val)
                s = str(val).strip().lower()
                m = re.search(r"(\d+(?:[\.,]\d+)?)\s*%", s)
                if m:
                    try:
                        num = float(m.group(1).replace(',', '.'))
                        return num > 80.0
                    except Exception:
                        pass
                positives = ['si', 'sí', 'cumple', 'ok', 'yes', 'true', 'cumplido']
                negatives = ['no cumple', 'no', 'false', 'falso']
                for neg in negatives:
                    if neg in s:
                        return False
                return any(x in s for x in positives)
            
            technical_summary.append({
                'Proveedor': p['name'],
                'Bomba': summary.get('Bomba', 'N/A'),
                'Cumplimiento': '✅ CUMPLE' if _is_cumple(compliance) else '❌ NO CUMPLE',
                'Status': _is_cumple(compliance)
            })
        
        if technical_summary:
            df_tech = pd.DataFrame(technical_summary)
            
            # Métricas
            col1, col2, col3 = st.columns(3)
            cumple_count = sum(1 for t in technical_summary if t['Status'])
            total = len(technical_summary)
            cumple_pct = (cumple_count/total*100) if total > 0 else 0
            
            col1.metric("✅ Cumplen", f"{cumple_count}/{total}")
            col2.metric("📊 Porcentaje", f"{cumple_pct:.1f}%")
            col3.metric("❌ No Cumplen", f"{total - cumple_count}/{total}")
            
            # Tabla estilizada
            def style_tech_table(row):
                styles = [''] * len(row)
                if '✅' in str(row['Cumplimiento']):
                    styles[row.index.get_loc('Cumplimiento')] = 'background-color: #d4edda; font-weight: bold; color: #155724'
                elif '❌' in str(row['Cumplimiento']):
                    styles[row.index.get_loc('Cumplimiento')] = 'background-color: #f8d7da; font-weight: bold; color: #721c24'
                return styles
            
            styled = df_tech[['Proveedor', 'Bomba', 'Cumplimiento']].style.apply(style_tech_table, axis=1)
            st.dataframe(styled, use_container_width=True, hide_index=True)
        else:
            show_no_data_message("Sin Datos Técnicos", "No hay evaluaciones técnicas disponibles")
    
    with tab3:
        st.markdown("#### Análisis de Costos Energéticos (TLCC)")
        
        energy_data = []
        for p in processed:
            energy_data.append({
                'Proveedor': p['name'],
                'PMM 4½"': p['TLCC PMM_4_5']['totalCost'],
                'PMM 3½"': p['TLCC PMM_3_5']['totalCost'],
                'AM 4½"': p['am_4_5']['totalCost'],
                'AM 3½"': p['am_3_5']['totalCost']
            })
        
        if energy_data:
            df_energy = pd.DataFrame(energy_data)
            df_energy['Promedio'] = df_energy[['PMM 4½"', 'PMM 3½"', 'AM 4½"', 'AM 3½"']].replace(0, np.nan).mean(axis=1)
            df_energy_sorted = df_energy.dropna(subset=['Promedio']).sort_values('Promedio')
            
            if not df_energy_sorted.empty:
                best_energy = df_energy_sorted.iloc[0]
                st.success(f"🏆 **Mejor Eficiencia Energética**: {best_energy['Proveedor']} (Promedio: ${best_energy['Promedio']:,.2f})")
                
                # Gráfico de barras horizontales
                fig = go.Figure()
                
                for col in ['PMM 4½"', 'PMM 3½"', 'AM 4½"', 'AM 3½"']:
                    fig.add_trace(go.Bar(
                        name=col,
                        y=df_energy_sorted['Proveedor'],
                        x=df_energy_sorted[col],
                        orientation='h'
                    ))
                
                fig.update_layout(
                    barmode='group',
                    title='Comparativa de Costos Energéticos por Proveedor',
                    xaxis_title='Costo Total (USD)',
                    yaxis_title='Proveedor',
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='white'),
                    height=400,
                    legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1)
                )
                
                st.plotly_chart(fig, use_container_width=True)
            else:
                show_no_data_message("Sin Datos Válidos", "No hay costos energéticos para comparar")
        else:
            show_no_data_message("Sin Datos Energéticos", "No se encontraron datos TLCC")
    
    st.markdown("---")
    
    # ===== RECOMENDACIÓN FINAL =====
    st.markdown("""
    <div style='background: linear-gradient(135deg, #FF6A00 0%, #f5576c 100%);
                padding: 2rem; border-radius: 15px; margin-top: 2rem;
                box-shadow: 0 10px 40px rgba(245, 87, 108, 0.4);'>
        <h2 style='color: white; text-align: center; margin-bottom: 1.5rem;'>
            💡 RECOMENDACIÓN EJECUTIVA
        </h2>
    """, unsafe_allow_html=True)
    
    if ranked_providers:
        winner_name, winner_data = ranked_providers[0]
        # Acotar y formatear para presentación
        econ_disp = min(max(float(winner_data.get('economico', 0)), 0.0), 40.0)
        tech_disp = min(max(float(winner_data.get('tecnico', 0)), 0.0), 35.0)
        energy_disp = min(max(float(winner_data.get('energetico', 0)), 0.0), 25.0)
        total_disp = min(max(float(winner_data.get('total', 0)), 0.0), 100.0)

        st.markdown(f"""
        <div style='color: white; font-size: 1.05rem; line-height: 1.6;'>
            <p><strong>Basado en el análisis integral de {len(processed)} proveedores</strong>, se recomienda seleccionar a
            <strong style='font-size: 1.25rem; color: #FFE66D;'>{winner_name}</strong> como proveedor principal.</p>
        </div>
        """, unsafe_allow_html=True)

        # Mostrar métricas con componentes nativos de Streamlit para mejor presentación
        m1, m2, m3, m4 = st.columns(4)
        m1.metric(label="PUNTAJE TOTAL", value=f"{total_disp:.1f} / 100")
        m2.metric(label="ECONÓMICO", value=f"{econ_disp:.1f} / 40")
        m3.metric(label="TÉCNICO", value=f"{tech_disp:.1f} / 35")
        m4.metric(label="ENERGÉTICO", value=f"{energy_disp:.1f} / 25")

        st.caption("Esta recomendación considera criterios económicos, técnicos y energéticos de forma equilibrada.")

        st.markdown("</div>", unsafe_allow_html=True)
        # Filtrar solo los que tienen datos
        df_pmm = pd.DataFrame(tlcc_pmm_data)
        df_pmm_valid = df_pmm[(df_pmm['4½" Total'] > 0) | (df_pmm['3½" Total'] > 0)]

        if not df_pmm_valid.empty:
            # Formatear valores
            for col in ['4½" Total', '4½" Fijo', '3½" Total', '3½" Fijo']:
                df_pmm_valid[col] = df_pmm_valid[col].apply(lambda x: f"${x:,.2f}" if x > 0 else "—")

            st.dataframe(df_pmm_valid, use_container_width=True, hide_index=True)

            # Encontrar mejores ofertas
            col1, col2 = st.columns(2)
            df_pmm_num = pd.DataFrame(tlcc_pmm_data)

            with col1:
                if df_pmm_num[df_pmm_num['4½" Total'] > 0].shape[0] > 0:
                    best_45 = df_pmm_num[df_pmm_num['4½" Total'] > 0].sort_values('4½" Total').iloc[0]
                    best_45_val = best_45['4½" Total']
                    st.success(f'🏆 **Mejor 4½"**: {best_45["Proveedor"]} → ${best_45_val:,.2f}')
                else:
                    show_no_data_message("Sin Propuesta", "No hay ofertas para 4½\"")

            with col2:
                if df_pmm_num[df_pmm_num['3½" Total'] > 0].shape[0] > 0:
                    best_35 = df_pmm_num[df_pmm_num['3½" Total'] > 0].sort_values('3½" Total').iloc[0]
                    best_35_val = best_35['3½" Total']
                    st.success(f'🏆 **Mejor 3½"**: {best_35["Proveedor"]} → ${best_35_val:,.2f}')
                else:
                    show_no_data_message("Sin Propuesta", "No hay ofertas para 3½\"")
        else:
            show_no_data_message("Sin Datos TLCC PMM", "No hay costos energéticos disponibles")

        with tab2:
            df_am = pd.DataFrame(tlcc_am_data)
            df_am_valid = df_am[(df_am['4½" Total'] > 0) | (df_am['3½" Total'] > 0)]

            if not df_am_valid.empty:
                for col in ['4½" Total', '4½" Fijo', '3½" Total', '3½" Fijo']:
                    df_am_valid[col] = df_am_valid[col].apply(lambda x: f"${x:,.2f}" if x > 0 else "—")

                st.dataframe(df_am_valid, use_container_width=True, hide_index=True)

                col1, col2 = st.columns(2)
                df_am_num = pd.DataFrame(tlcc_am_data)

                with col1:
                    if df_am_num[df_am_num['4½" Total'] > 0].shape[0] > 0:
                        best_45 = df_am_num[df_am_num['4½" Total'] > 0].sort_values('4½" Total').iloc[0]
                        best_45_val = best_45['4½" Total']
                        st.success(f'🏆 **Mejor 4½"**: {best_45["Proveedor"]} → ${best_45_val:,.2f}')
                    else:
                        show_no_data_message("Sin Propuesta", "No hay ofertas para 4½\"")

                with col2:
                    if df_am_num[df_am_num['3½" Total'] > 0].shape[0] > 0:
                        best_35 = df_am_num[df_am_num['3½" Total'] > 0].sort_values('3½" Total').iloc[0]
                        best_35_val = best_35['3½" Total']
                        st.success(f'🏆 **Mejor 3½"**: {best_35["Proveedor"]} → ${best_35_val:,.2f}')
                    else:
                        show_no_data_message("Sin Propuesta", "No hay ofertas para 3½\"")
            else:
                show_no_data_message("Sin Datos TLCC AM", "No hay costos energéticos disponibles")

    st.markdown("---")
# ==================== SECCIÓN COSTOS ====================
elif motor_choice == 'COSTOS':
    st.markdown('<div class="section-header">💰 Ranking Económico por Categoría y Tamaño de Tubería</div>', unsafe_allow_html=True)

    # Añadir selector para ver tablas AM o PMM (mapeo provisto por el usuario)
    view_choice = st.radio('Ver tablas', ['AM', 'PMM'], index=0, horizontal=True, help='Selecciona si quieres ver las tablas AM o PMM')

    ROW_MAP = {
        'NOMBRE BOMBA': 2,
        'EQUIPO DE SUPERFICIE': 3,
        'EQUIPO ESP': 4,
        'Cable + Accesorios': 5,
        'Y-TOOL': 6,
        'Servicios': 8,
        'EQUIPO + SERVICIO': 9,
        'TOTAL': 10,
        'TLCC': 11,
        'EQUIPO+SERVICIOS+TLCC': 12
    }

    raw_data = []
    # Mapeo de tablas por vista (según lo indicado):
    # AM: tabla 3 -> AM 4-1/2", tabla 4 -> AM 3-1/2"
    # PMM: tabla 2 -> PMM 4-1/2", tabla 1 -> PMM 3-1/2"
    allowed_tables = {
        'AM': [3, 4],
        'PMM': [2, 1]
    }

    import re
    for p in processed:
        filename = p['name']
        for t in p.get('fronteraTables', []):
            data = t['data']
            if not data or len(data) < 13: continue

            # Determinar número de tabla a partir del título (p.ej. 'Tabla 1: ...')
            title_lower = (t.get('title') or '').lower()
            m = re.search(r'tabla\s*(\d+)', title_lower)
            table_num = int(m.group(1)) if m else None
            if table_num is None or table_num not in allowed_tables.get(view_choice, []):
                # ignorar tablas que no correspondan a la vista seleccionada
                continue

            provider = t['providerName']
            pump = t['pumpName']
            pipe = t['pipeSize']

            # Buscar dinámicamente la fila que contiene 'EQUIPO+SERVICIOS+TLCC'
            total_idx = None
            for i, row in enumerate(data):
                if len(row) >= 1 and 'equipo+servicios+tlcc' in str(row[0]).lower():
                    total_idx = i
                    break
            if total_idx is None:
                # no se encontró la fila de totales; saltar
                continue

            def clean_num(val):
                if pd.isna(val): return 0.0
                s = str(val).replace("$", "").replace("USD", "").replace(",", "").strip()
                s = re.sub(r"[^0-9.\-]", "", s)
                try:
                    return float(s) if s not in ['', None] else 0.0
                except:
                    return 0.0

            full_total = clean_num(data[total_idx][1]) if len(data[total_idx]) > 1 else 0.0
            rrigo_total = clean_num(data[total_idx][2]) if len(data[total_idx]) > 2 else 0.0
            alt_total = clean_num(data[total_idx][3]) if len(data[total_idx]) > 3 else 0.0

            if full_total > 0:
                raw_data.append({'cat': 'FULL PRICE', 'pipe': pipe, 'prov': provider, 'pump': pump, 'filename': filename, 'col': 1, 'data': data})
            if rrigo_total > 0:
                raw_data.append({'cat': 'R-R-I-G-O', 'pipe': pipe, 'prov': provider, 'pump': pump, 'filename': filename, 'col': 2, 'data': data})
            if alt_total > 0:
                raw_data.append({'cat': 'ALTERNATIVA AHORRO', 'pipe': pipe, 'prov': provider, 'pump': pump, 'filename': filename, 'col': 3, 'data': data})

    categories = ['FULL PRICE', 'R-R-I-G-O', 'ALTERNATIVA AHORRO']
    pipes = ['3-1/2"', '4-1/2"']
    
    category_icons = {
        'FULL PRICE': '💎',
        'R-R-I-G-O': '🔄',
        'ALTERNATIVA AHORRO': '💰'
    }

    for cat in categories:
        st.markdown(f"### {category_icons.get(cat, '📊')} {cat}")
        cols = st.columns(2)
        col_idx = 0

        for pipe in pipes:
            with cols[col_idx]:
                st.markdown(f"**🔧 Tubería {pipe}**")
                filtered = [d for d in raw_data if d['cat'] == cat and d['pipe'] == pipe]
                
                if not filtered:
                    show_no_data_message("Sin Ofertas", f"No hay propuestas para {cat} - {pipe}")
                    col_idx += 1
                    continue

                table_data = []
                for var_name, row_idx in ROW_MAP.items():
                    row = {'VARIABLE': var_name}
                    for item in filtered:
                        prov = item['prov']
                        val = item['data'][row_idx][item['col']] if row_idx < len(item['data']) and item['col'] < len(item['data'][row_idx]) else 0.0
                        s = str(val).replace(", ", "").replace(",", "").replace("USD", "").strip()
                        try:
                            num = float(s)
                            row[prov] = f"${num:,.0f}" if num > 0 else "—"
                        except:
                            row[prov] = str(val).strip() if val else "—"
                    table_data.append(row)

                df_table = pd.DataFrame(table_data)
                if len(df_table) > 0 and len(df_table.columns) > 1:
                    df_display = df_table.copy()
                    df_display['VARIABLE'] = df_display['VARIABLE'].str.replace(' + ', ' ', regex=True)

                    min_col = None
                    if 'EQUIPO+SERVICIOS+TLCC' in df_display['VARIABLE'].values:
                        total_row = df_display[df_display['VARIABLE'] == 'EQUIPO+SERVICIOS+TLCC']
                        if not total_row.empty:
                            numeric_values = []
                            for col in df_display.columns[1:]:
                                try:
                                    val = float(str(total_row[col].iloc[0]).replace(", ", "").replace(",", ""))
                                    numeric_values.append((col, val))
                                except:
                                    numeric_values.append((col, float('inf')))
                            if numeric_values:
                                min_col = min(numeric_values, key=lambda x: x[1])[0]

                    def highlight_winner(row):
                        styles = [''] * len(row)
                        if row['VARIABLE'] == 'EQUIPO+SERVICIOS+TLCC' and min_col and min_col in row.index:
                            idx = row.index.get_loc(min_col)
                            styles[idx] = 'background-color: #d4edda; font-weight: bold; color: #155724'
                        return styles

                    styled_df = df_display.style.apply(highlight_winner, axis=1)
                    st.dataframe(styled_df, use_container_width=True, hide_index=True)
                else:
                    st.dataframe(df_table, use_container_width=True, hide_index=True)
            col_idx += 1
            if col_idx >= 2:
                col_idx = 0
        st.markdown("---")

    # Ganadores
    full_offers = []
    rrigo_offers = []
    alt_offers = []

    import re
    for p in processed:
        for t in p.get('fronteraTables', []):
            data = t['data']
            if not data or len(data) < 13: continue

            # Filtrar según la selección AM/PMM (usar mismas tablas permitidas)
            title_lower = (t.get('title') or '').lower()
            m = re.search(r'tabla\s*(\d+)', title_lower)
            table_num = int(m.group(1)) if m else None
            if table_num is None or table_num not in allowed_tables.get(view_choice, []):
                continue

            provider = t['providerName']
            pump = t['pumpName']
            pipe = t['pipeSize']
            title = f"{provider} - {pump}"

            total_row = None
            for row in data:
                if len(row) >= 1 and "equipo+servicios+tlcc" in str(row[0]).lower():
                    total_row = row
                    break
            if not total_row or len(total_row) < 4: continue

            def clean_val(val):
                if pd.isna(val): return 0.0
                s = str(val).replace(", ", "").replace(",", "").replace("USD", "").strip()
                try: return float(s)
                except: return 0.0

            full = clean_val(total_row[1])
            rrigo = clean_val(total_row[2])
            alt = clean_val(total_row[3])

            offer = {
                'title': title,
                'provider': provider,
                'pump': pump,
                'pipe': pipe,
                'filename': p['name']
            }

            if full > 0:
                full_offers.append({**offer, 'value': full})
            if rrigo > 0:
                rrigo_offers.append({**offer, 'value': rrigo})
            if alt > 0:
                alt_offers.append({**offer, 'value': alt})

    def get_best_per_pipe(df):
        # Retorna un dict: { pipe: best_row }
        out = {}
        if df.empty:
            return out
        for pipe_val in df['pipe'].unique():
            df_pipe = df[df['pipe'] == pipe_val]
            if not df_pipe.empty:
                best_row = df_pipe.sort_values('value').iloc[0]
                out[pipe_val] = best_row
        return out

    st.markdown('<div class="section-header">🏆 Ganadores por Categoría y Tubería</div>', unsafe_allow_html=True)
    cols = st.columns(3)
    winners = {}

    with cols[0]:
        st.markdown("**💎 FULL PRICE**")
        df_full = pd.DataFrame(full_offers)
        if not df_full.empty:
            best_map = get_best_per_pipe(df_full)
            winners['FULL PRICE'] = best_map
            for pipe in pipes:
                st.markdown(f"**🔧 Tubería {pipe}**")
                best = best_map.get(pipe)
                if not best is None:
                    st.success(f"**{pipe}**: {best['title']} → **${best['value']:,.0f}**")
                else:
                    show_no_data_message(f"Sin Propuesta {pipe}", "No hay ofertas")
        else:
            show_no_data_message("Sin Ofertas", "No hay propuestas FULL PRICE")

    with cols[1]:
        st.markdown("**🔄 R-R-I-G-O**")
        df_rrigo = pd.DataFrame(rrigo_offers)
        if not df_rrigo.empty:
            best_map = get_best_per_pipe(df_rrigo)
            winners['R-R-I-G-O'] = best_map
            for pipe in pipes:
                st.markdown(f"**🔧 Tubería {pipe}**")
                best = best_map.get(pipe)
                if not best is None:
                    st.success(f"**{pipe}**: {best['title']} → **${best['value']:,.0f}**")
                else:
                    show_no_data_message(f"Sin Propuesta {pipe}", "No hay ofertas")
        else:
            show_no_data_message("Sin Ofertas", "No hay propuestas R-R-I-G-O")

    with cols[2]:
        st.markdown("**💰 ALTERNATIVA AHORRO**")
        df_alt = pd.DataFrame(alt_offers)
        if not df_alt.empty:
            best_map = get_best_per_pipe(df_alt)
            winners['ALTERNATIVA AHORRO'] = best_map
            for pipe in pipes:
                st.markdown(f"**🔧 Tubería {pipe}**")
                best = best_map.get(pipe)
                if not best is None:
                    st.success(f"**{pipe}**: {best['title']} → **${best['value']:,.0f}**")
                else:
                    show_no_data_message(f"Sin Propuesta {pipe}", "No hay ofertas")
        else:
            show_no_data_message("Sin Ofertas", "No hay propuestas ALTERNATIVA AHORRO")

    st.success("✅ Análisis Económico Completado")

# ==================== SECCIÓN EV TÉCNICA ====================
elif motor_choice == 'EV TECNICA':
    st.markdown('<div class="section-header">🔬 EVALUACIÓN TÉCNICA DE DISEÑOS</div>', unsafe_allow_html=True)
    comp_minimo = []
    comp_optimo = []
    comp_maximo = []

    def _is_cumple(val):
        import re
        if val is None:
            return False
        if isinstance(val, (int, float)):
            return bool(val)
        s = str(val).strip().lower()
        m = re.search(r"(\d+(?:[\.,]\d+)?)\s*%", s)
        if m:
            try:
                num = float(m.group(1).replace(',', '.'))
                return num > 80.0
            except Exception:
                pass
        positives = ['si', 'sí', 'cumple', 'ok', 'yes', 'true', 'cumplido']
        negatives = ['no cumple', 'no', 'false', 'falso']
        for neg in negatives:
            if neg in s:
                return False
        return any(x in s for x in positives)

    def color_compliance_cell(row):
        styles = [''] * len(row) 
        col_name = 'Parámetro de Diseño Solicitado'
        try:
            col_index = row.index.get_loc(col_name)
            status = row[col_name]
        except KeyError:
            return styles 

        if status == 'CUMPLE':
            styles[col_index] = 'background-color: #d4edda; font-weight: bold; color: #155724; border-radius: 5px; padding: 5px;' 
        elif status == 'NO CUMPLE':
            styles[col_index] = 'background-color: #f8d7da; font-weight: bold; color: #721c24; border-radius: 5px; padding: 5px;'
        return styles

    def color_bfpd_score_cell(row):
        styles = [''] * len(row)
        col_name = 'BFPD Score'
        try:
            col_index = row.index.get_loc(col_name)
            score = row[col_name]
        except KeyError:
            return styles

        try:
            val = float(score)
        except Exception:
            return styles

        if val >= 100:
            styles[col_index] = 'background-color: #d4edda; font-weight: bold; color: #155724; padding: 5px;'
        elif val >= 80:
            styles[col_index] = 'background-color: #fff3cd; font-weight: bold; color: #856404; padding: 5px;'
        else:
            styles[col_index] = 'background-color: #f8d7da; font-weight: bold; color: #721c24; padding: 5px;'
        return styles

    def make_color_closest_bfpd(best_index):
        def _color(row):
            styles = [''] * len(row)
            if best_index is None:
                return styles
            try:
                col_index = row.index.get_loc('BFPD')
            except Exception:
                return styles
            try:
                if row.name == best_index:
                    styles[col_index] = 'background-color: #d4edda; font-weight: bold; color: #155724; padding: 5px;'
            except Exception:
                return styles
            return styles
        return _color

    has_technical_data = False
    
    for p in processed:
        diseno = p.get('disenoData') or {}
        summary = diseno.get('summary', {})
        km = diseno.get('key_metrics', {})

        # Si hay datos técnicos
        if summary or km:
            has_technical_data = True

        compliance_final_d81 = p.get('compliance_final_d81')
        if compliance_final_d81 is None:
            compliance_final_d81 = (p.get('oferta_meta') or {}).get('D81')
        if compliance_final_d81 is None:
            compliance_final_d81 = summary.get('Compliance_Final_D81')
        if compliance_final_d81 is None:
            compliance_final_d81 = 'N/A'
        
        final_status_is_compliant = _is_cumple(compliance_final_d81)
        final_status_text = 'CUMPLE' if final_status_is_compliant else 'NO CUMPLE'

        with st.expander(f"**📄 {p['name']} - {summary.get('Proveedor', 'N/A')}**", expanded=False):
            col1, col2 = st.columns([1, 1])

            with col1:
                st.markdown("### 📋 Resumen")
                if summary:
                    st.markdown(f"""
                    - **Pozo:** `{summary.get('Pozo', 'N/A')}`
                    - **Bomba:** `{summary.get('Bomba', 'N/A')}`
                    - **Cuerpos:** `{summary.get('Cuerpos de Bomba', 'N/A')}`
                    - **Etapas:** `{summary.get('Numero de Etapas', 'N/A')}`
                    """)
                else:
                    show_no_data_message("Sin Datos", "No hay información de resumen disponible")

                st.markdown("### 📊 Parámetros Clave")
                if km and any(km.values()):
                    cases = ['Mínimo', 'Óptimo', 'Máximo']
                    params = ['TDH', 'BFPD','Frecuencia', 'kva', 'PIP', 'P estatica', 'P descarga', 
                             'Nivel de fluido, MD', 'Eficiencia de la bomba', 'Eficiencia del motor', 
                             'Shaft load Pump', 'Carga del motor HP', 'Carga del motor Amp', 'Carga de la zapata']
                    try:
                        data_for_df = []
                        for param in params:
                            row = {'Parámetro': param}
                            for c in cases:
                                row[c] = km.get(c, {}).get(param, 'N/A')
                            data_for_df.append(row)
                            
                        df_key = pd.DataFrame(data_for_df)
                        st.dataframe(df_key, use_container_width=True, hide_index=True)
                    except Exception as e:
                        show_no_data_message("Error en Datos", f"No se pudieron cargar las métricas: {e}")
                else:
                    show_no_data_message("Sin Métricas", "No hay parámetros clave disponibles")

            with col2:
                st.markdown("### ✅ Parámetros de Diseño")
                oferta = p.get('oferta') or diseno.get('oferta') or []
                
                if not oferta:
                    show_no_data_message("Sin Datos de Oferta", "No se encontró información en OFERTA ECONOMICA")
                else:
                    df_of = pd.DataFrame(oferta)
                    display_df = df_of.rename(columns={'title': 'Título', 'valor': 'Valor', 'cumple': 'Cumple'})
                    st.dataframe(display_df, use_container_width=True, hide_index=True)

                st.markdown("---")
                st.markdown("### 🎯 Cumplimiento General")
                
                if compliance_final_d81 == 'N/A':
                    show_no_data_message("Sin Evaluación", "No se encontró el estado de cumplimiento final (D81)")
                elif final_status_is_compliant:
                    st.success(f'✅ **CUMPLIMIENTO GENERAL:** {compliance_final_d81}')
                else:
                    st.error(f'❌ **CUMPLIMIENTO GENERAL:** {compliance_final_d81}')

        base = {
            'Archivo': p['name'],
            'Proveedor': summary.get('Proveedor', 'N/A'),
            'Pozo': summary.get('Pozo', 'N/A'),
            'Bomba': summary.get('Bomba', 'N/A'),
            'Cuerpos de Bomba': summary.get('Cuerpos de Bomba', 'N/A'),
            'Numero de Etapas': summary.get('Numero de Etapas', 'N/A'),
            'Tiempo de entrega': summary.get('Tiempo de entrega', 'N/A'),
            'Parámetro de Diseño Solicitado': final_status_text 
        }
        
        for case in ['Mínimo', 'Óptimo', 'Máximo']:
            row = base.copy()
            row.update(km.get(case, {}))
            if case == 'Mínimo':
                comp_minimo.append(row)
            elif case == 'Óptimo':
                comp_optimo.append(row)
            else:
                comp_maximo.append(row)

    if not has_technical_data:
        show_no_data_message("Sin Evaluación Técnica", "No se encontraron datos técnicos en ningún archivo cargado")
        st.info("💡 Asegúrate de que los archivos Excel contengan la hoja 'DISEÑO' con la información requerida")
    else:
        st.success("✅ Evaluación Técnica Completada")

    st.markdown('<div class="section-header">📊 COMPARATIVA DE DISEÑOS</div>', unsafe_allow_html=True)
    
    cols_order = [
        'Archivo', 'Proveedor', 'Pozo', 'Bomba', 'Cuerpos de Bomba', 'Numero de Etapas', 
        'TDH', 'BFPD', 'Frecuencia', 'Tiempo de entrega',
        'Parámetro de Diseño Solicitado'
    ]

    case_icons = {
        'MÍNIMO': '📉',
        'ÓPTIMO': '🎯',
        'MÁXIMO': '📈'
    }

    # Buscar valores SLA globales (si existen) en los archivos procesados
    sla_global = None
    for p in processed:
        try:
            s = (p.get('oferta_meta') or {}).get('SLA')
            if s:
                sla_global = s
                break
        except Exception:
            continue

    for name, data in [('MÍNIMO', comp_minimo), ('ÓPTIMO', comp_optimo), ('MÁXIMO', comp_maximo)]:
        if data:
            # Mapear caso a campo SLA
            map_field = {'MÍNIMO': 'BFPD_min', 'ÓPTIMO': 'BFPD_obj', 'MÁXIMO': 'BFPD_max'}
            sla_value = None
            if sla_global:
                sla_value = sla_global.get(map_field.get(name))

            # Mostrar título con el valor objetivo cuando exista (al lado del título)
            if sla_value is not None:
                try:
                    sla_disp = float(sla_value)
                    sla_formatted = f"{sla_disp:,.2f}"
                except Exception:
                    sla_formatted = str(sla_value)
                # Etiqueta según el caso (mínimo / objetivo / máximo)
                label_map = {'MÍNIMO': 'Min', 'ÓPTIMO': 'Obj', 'MÁXIMO': 'Max'}
                lbl = label_map.get(name, 'Obj')
                st.markdown(f"### {case_icons.get(name, '📊')} CASO {name} : {sla_formatted} BFPD ({lbl})")
            else:
                st.markdown(f"### {case_icons.get(name, '📊')} CASO {name}")

            df = pd.DataFrame(data)
            df = df.reindex(columns=[c for c in cols_order if c in df.columns])

            # Resaltar el valor de BFPD Diseño más cercano al objetivo (si existe)
            closest_archivo = None
            try:
                if 'BFPD' in df.columns and sla_value is not None:
                    df_tmp = df.copy()
                    df_tmp['_bfpd_num'] = pd.to_numeric(df_tmp['BFPD'], errors='coerce')
                    df_valid = df_tmp.dropna(subset=['_bfpd_num'])
                    if not df_valid.empty:
                        sla_f = float(sla_value)
                        idx = (df_valid['_bfpd_num'] - sla_f).abs().idxmin()
                        closest_archivo = df_tmp.loc[idx, 'Archivo']
            except Exception:
                closest_archivo = None

            def color_closest_bfpd(row):
                styles = [''] * len(row)
                if 'BFPD' not in row.index:
                    return styles
                try:
                    col_idx = row.index.get_loc('BFPD')
                except Exception:
                    return styles
                try:
                    if closest_archivo is not None and row.get('Archivo') == closest_archivo:
                        styles[col_idx] = 'background-color: #cce5ff; font-weight: bold; color: #084298; padding: 5px;'
                except Exception:
                    pass
                return styles

            styled_df = df.style.apply(color_compliance_cell, axis=1).apply(color_closest_bfpd, axis=1)
            st.dataframe(styled_df, use_container_width=True)
        else:
            show_no_data_message(f"Sin Datos - Caso {name}", "No hay información comparativa disponible")

# ==================== SECCIÓN TLCC ====================
elif motor_choice in ['TLCC PMM', 'TLCC AM']:
    cfg = COLUMN_INDICES[motor_choice]
    key45 = cfg['4_5']['key']
    key35 = cfg['3_5']['key']
    
    st.markdown(f'<div class="section-header">⚙️ Resultados para {motor_choice}</div>', unsafe_allow_html=True)

    valid_processed = [p for p in processed if p[key45]['totalCost'] > 0 or p[key35]['totalCost'] > 0]
    if not valid_processed:
        show_no_data_message("Sin Ofertas Válidas", f"Ningún proveedor presentó oferta válida para {motor_choice}")
        st.info("💡 Verifica que los archivos Excel contengan datos en las columnas correspondientes de la hoja TLCC")
        st.stop()

    summary_list = []
    for p in valid_processed:
        summary_list.append({
            'name': p['name'],
            'total45': p[key45]['totalCost'],
            'total35': p[key35]['totalCost'],
            'fixed45': p[f'{key45}_fixed'],
            'fixed35': p[f'{key35}_fixed']
        })
    df_summary = pd.DataFrame(summary_list)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"### 🏆 Mejor {cfg['4_5']['title']}")
        valid45 = df_summary[df_summary['total45'] > 0]
        if not valid45.empty:
            best = valid45.sort_values('total45').iloc[0]
            st.success(f"**{best['name']}** — ${best['total45']:,.2f}")
        else:
            show_no_data_message("Sin Propuesta", f"No hay ofertas para {cfg['4_5']['title']}")
    
    with col2:
        st.markdown(f"### 🏆 Mejor {cfg['3_5']['title']}")
        valid35 = df_summary[df_summary['total35'] > 0]
        if not valid35.empty:
            best = valid35.sort_values('total35').iloc[0]
            st.success(f"**{best['name']}** — ${best['total35']:,.2f}")
        else:
            show_no_data_message("Sin Propuesta", f"No hay ofertas para {cfg['3_5']['title']}")

    st.markdown("---")
    st.markdown("### 📊 Tabla Comparativa")
    display_cols = ['name']
    if df_summary['total45'].gt(0).any(): display_cols.append('total45')
    if df_summary['total35'].gt(0).any(): display_cols.append('total35')
    
    if len(display_cols) > 1:
        st.dataframe(
            df_summary[display_cols].rename(columns={
                'name': 'Proveedor', 
                'total45': 'Total 4½"', 
                'total35': 'Total 3½"'
            }), 
            use_container_width=True,
            hide_index=True
        )
    else:
        show_no_data_message("Sin Datos Comparativos", "No hay suficientes datos para generar la tabla comparativa")

    # GRÁFICOS MEJORADOS CON PLOTLY
    meses = list(range(1, 61))
    
    if df_summary['total45'].gt(0).any():
        st.markdown(f'### 📈 Distribución Mensual - {cfg["4_5"]["title"]}')
        
        rows = []
        for p in valid_processed:
            if p[key45]['totalCost'] <= 0: continue
            for m, v in zip(meses, p[key45]['rawValues']):
                rows.append({'Proveedor': p['name'], 'Mes': m, 'Costo Acumulado (USD)': v})
        
        if rows:
            df_line = pd.DataFrame(rows)
            
            fig = go.Figure()
            colors = ['#667eea', '#764ba2', '#f093fb', '#4facfe', '#43e97b']
            
            for i, proveedor in enumerate(df_line['Proveedor'].unique()):
                df_prov = df_line[df_line['Proveedor'] == proveedor]
                fig.add_trace(go.Scatter(
                    x=df_prov['Mes'],
                    y=df_prov['Costo Acumulado (USD)'],
                    mode='lines+markers',
                    name=proveedor,
                    line=dict(width=3, color=colors[i % len(colors)]),
                    marker=dict(size=6, symbol='circle'),
                    hovertemplate='<b>%{fullData.name}</b><br>Mes: %{x}<br>Costo: $%{y:,.0f}<extra></extra>'
                ))
            
            fig.update_layout(
                title=dict(text=f'Evolución de Costos - {cfg["4_5"]["title"]}', font=dict(size=20, color=COLOR_FUENTE)),
                xaxis=dict(title='Mes', showgrid=True, gridcolor='rgba(255,255,255,0.1)', dtick=5),
                yaxis=dict(title='Costo Acumulado (USD)', showgrid=True, gridcolor='rgba(255,255,255,0.1)', tickformat='$,.0f'),
                hovermode='x unified',
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='white'),
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                height=500
            )
            
            st.plotly_chart(fig, use_container_width=True)
        else:
            show_no_data_message("Sin Datos de Gráfico", "No hay valores mensuales disponibles para graficar")
    else:
        show_no_data_message(f"Sin Datos {cfg['4_5']['title']}", "No hay ofertas disponibles para este tamaño")

    if df_summary['total35'].gt(0).any():
        st.markdown(f'### 📈 Distribución Mensual - {cfg["3_5"]["title"]}')
        
        rows = []
        for p in valid_processed:
            if p[key35]['totalCost'] <= 0: continue
            for m, v in zip(meses, p[key35]['rawValues']):
                rows.append({'Proveedor': p['name'], 'Mes': m, 'Costo Acumulado (USD)': v})
        
        if rows:
            df_line = pd.DataFrame(rows)
            
            fig = go.Figure()
            colors = get_color_sequence()
            
            for i, proveedor in enumerate(df_line['Proveedor'].unique()):
                df_prov = df_line[df_line['Proveedor'] == proveedor]
                fig.add_trace(go.Scatter(
                    x=df_prov['Mes'],
                    y=df_prov['Costo Acumulado (USD)'],
                    mode='lines+markers',
                    name=proveedor,
                    line=dict(width=3, color=colors[i % len(colors)]),
                    marker=dict(size=6, symbol='circle'),
                    hovertemplate='<b>%{fullData.name}</b><br>Mes: %{x}<br>Costo: $%{y:,.0f}<extra></extra>'
                ))
            
            fig.update_layout(
                title=dict(text=f'Evolución de Costos - {cfg["3_5"]["title"]}', font=dict(size=20, color=COLOR_FUENTE)),
                xaxis=dict(title='Mes', showgrid=True, gridcolor='rgba(255,255,255,0.1)', dtick=5),
                yaxis=dict(title='Costo Acumulado (USD)', showgrid=True, gridcolor='rgba(255,255,255,0.1)', tickformat='$,.0f'),
                hovermode='x unified',
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='white'),
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                height=500
            )
            
            st.plotly_chart(fig, use_container_width=True)
        else:
            show_no_data_message("Sin Datos de Gráfico", "No hay valores mensuales disponibles para graficar")
    else:
        show_no_data_message(f"Sin Datos {cfg['3_5']['title']}", "No hay ofertas disponibles para este tamaño")

    st.markdown("---")
    st.markdown('<div class="section-header">💵 Valores Energéticos por Proveedor</div>', unsafe_allow_html=True)
    
    # Si COLOR_FUENTE es blanco/claro, la configuración es correcta.
    colf1, colf2 = st.columns(2)
    
    # --- GRÁFICA COLUMNA 1 (fixed45) ---
    with colf1:
        if df_summary['total45'].gt(0).any():
            st.markdown(f'### 💰 Costos TLCC - {cfg["4_5"]["title"]}') # Título ajustado para consistencia
            
            df_f = df_summary[df_summary['total45'] > 0][['name', 'fixed45']].sort_values('fixed45')
            df_f = df_f.rename(columns={'name': 'Proveedor', 'fixed45': 'Valor Fijo (USD)'})
            
            # Calcular el rango del Eje Y
            max_val = df_f['Valor Fijo (USD)'].max()
            y_range = [0, max_val * 1.15 if max_val > 0 else 1] # 15% de padding para el texto
            
            fig = go.Figure(data=[
                go.Bar(
                    x=df_f['Proveedor'],
                    y=df_f['Valor Fijo (USD)'],
                    text=df_f['Valor Fijo (USD)'].apply(lambda x: f'${x:,.0f}'),
                    textposition='outside',
                    marker=dict(
                        color=df_f['Valor Fijo (USD)'],
                        colorscale='Sunsetdark', # Cambiado a una escala de color más vibrante y oscura
                        line=dict(color='rgba(255, 255, 255, 0.5)', width=1) # Borde más claro para destacar
                    ),
                    hovertemplate='<b>%{x}</b><br>Valor Fijo: $%{y:,.0f}<extra></extra>'
                )
            ])
            
            fig.update_layout(
                title=dict(
                    text=f'Distribución de Costos TLCC - {cfg["4_5"]["title"]}', 
                    font=dict(size=20, color=COLOR_FUENTE)
                ),
                # Ajuste de Rango Y para que el texto no se corte
                yaxis=dict(
                    title='Valor Fijo (USD)', 
                    tickformat='$,.0f',
                    range=y_range 
                ),
                xaxis=dict(
                    title='Proveedor',
                    tickangle=-45, # Inclinar etiquetas para mejor lectura en espacio reducido
                    automargin=True # Asegura que el espacio se ajuste a las etiquetas
                ),
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color=COLOR_FUENTE), # Usar COLOR_FUENTE para todo el texto
                height=450, # Un poco más de altura para el texto inclinado
                margin=dict(t=50, b=100) # Asegurar márgenes suficientes
            )
            
            st.plotly_chart(fig, use_container_width=True)
        else:
            show_no_data_message(f"Sin Datos {cfg['4_5']['title']}", "No hay valores energéticos disponibles")

    # --- GRÁFICA COLUMNA 2 (fixed35) ---
    with colf2:
        if df_summary['total35'].gt(0).any():
            st.markdown(f'### 💰 Costos TLCC - {cfg["3_5"]["title"]}') # Título ajustado
            
            df_f = df_summary[df_summary['total35'] > 0][['name', 'fixed35']].sort_values('fixed35')
            df_f = df_f.rename(columns={'name': 'Proveedor', 'fixed35': 'Valor Fijo (USD)'})
            
            # Calcular el rango del Eje Y
            max_val = df_f['Valor Fijo (USD)'].max()
            y_range = [0, max_val * 1.15 if max_val > 0 else 1] # 15% de padding para el texto
            
            fig = go.Figure(data=[
                go.Bar(
                    x=df_f['Proveedor'],
                    y=df_f['Valor Fijo (USD)'],
                    text=df_f['Valor Fijo (USD)'].apply(lambda x: f'${x:,.0f}'),
                    textposition='outside',
                    marker=dict(
                        color=df_f['Valor Fijo (USD)'],
                        colorscale='Sunsetdark',
                        line=dict(color='rgba(255, 255, 255, 0.5)', width=1)
                    ),
                    hovertemplate='<b>%{x}</b><br>Valor Fijo: $%{y:,.0f}<extra></extra>'
                )
            ])
            
            fig.update_layout(
                title=dict(
                    text=f'Distribución de Costos TLCC - {cfg["3_5"]["title"]}', 
                    font=dict(size=20, color=COLOR_FUENTE)
                ),
                # Ajuste de Rango Y para que el texto no se corte
                yaxis=dict(
                    title='Valor Fijo (USD)', 
                    tickformat='$,.0f',
                    range=y_range
                ),
                xaxis=dict(
                    title='Proveedor',
                    tickangle=-45,
                    automargin=True
                ),
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color=COLOR_FUENTE),
                height=450,
                margin=dict(t=50, b=100)
            )
            
            st.plotly_chart(fig, use_container_width=True)
        else:
            show_no_data_message(f"Sin Datos {cfg['3_5']['title']}", "No hay valores energéticos disponibles")

    st.markdown("---")
    st.markdown("### 📂 Detalles por Proveedor")
    
    for p in valid_processed:
        with st.expander(f"📄 {p['name']}"):
            c1, c2 = st.columns(2)
            has_data = False
            
            if p[key45]['totalCost'] > 0:
                c1.metric(f"💵 Total {cfg['4_5']['title']}", f"${p[key45]['totalCost']:,.2f}")
                c1.metric(f"🔒 Fijo {cfg['4_5']['title']}", f"${p[f'{key45}_fixed']:,.2f}")
                has_data = True
            else:
                with c1:
                    show_no_data_message(f"Sin Datos {cfg['4_5']['title']}", "No hay propuesta")
            
            if p[key35]['totalCost'] > 0:
                c2.metric(f"💵 Total {cfg['3_5']['title']}", f"${p[key35]['totalCost']:,.2f}")
                c2.metric(f"🔒 Fijo {cfg['3_5']['title']}", f"${p[f'{key35}_fixed']:,.2f}")
                has_data = True
            else:
                with c2:
                    show_no_data_message(f"Sin Datos {cfg['3_5']['title']}", "No hay propuesta")
            
            if not has_data:
                st.warning("⚠️ Este proveedor no tiene datos válidos para ninguna configuración")

st.markdown("---")
st.markdown("""
<div style='text-align: center; opacity: 0.7; padding: 2rem;'>
    <p>Sistema de Evaluación ESP v1.0 | Desarrollado por AJM para análisis integral de ofertas</p>
</div>
""", unsafe_allow_html=True)
